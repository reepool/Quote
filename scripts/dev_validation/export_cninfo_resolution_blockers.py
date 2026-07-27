#!/usr/bin/env python3
"""Export current factor-blocking CNInfo corporate actions for operator review."""

from __future__ import annotations

import argparse
import json
import sqlite3
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


NO_EX_RIGHT_MARKERS = (
    "不除权",
    "不实施除权",
    "不进行除权",
    "不做除权",
    "不作除权",
    "不予除权",
    "不再对公司股票进行除权",
    "无需进行除权",
    "股票价格不作除权调整",
    "股票价格不进行除权调整",
    "不会对公司股票价格进行除权",
)
ASYMMETRIC_BENEFICIARY_MARKERS = (
    "不向原股东分配",
    "不向公司原股东分配",
    "不向股东分配",
    "流通股东",
    "非流通股东",
    "债权人",
    "重整投资人",
    "用于偿债",
    "用于清偿",
    "受让",
)
PRICE_ADJUSTMENT_MARKERS = (
    "除权参考价格的计算公式",
    "除权参考价将按照",
    "进行除权调整",
    "开盘参考价将进行调整",
)
EX_DATE_MARKERS = (
    "除权日为",
    "除权除息日为",
)
SPECIAL_CATEGORIES = {
    "股改分红",
    "重整转增",
    "承诺补偿",
    "偿债转增",
    "定向转增",
}


def _json_object(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not isinstance(value, str) or not value.strip():
        return {}
    try:
        parsed = json.loads(value)
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()[:10]
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _date_text(value: Any) -> str:
    parsed = _date(value)
    return parsed.isoformat() if parsed else ""


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _per_ten(value: Any) -> float | None:
    parsed = _number(value)
    return parsed * 10 if parsed is not None else None


def _compact(values: Iterable[Any], *, limit: int = 1200) -> str:
    text = "\n".join(
        str(value).strip()
        for value in values
        if value is not None and str(value).strip()
    )
    return text[:limit]


def _evidence_text(analysis: Mapping[str, Any]) -> str:
    return " ".join(
        str(item.get("exact_quote") or "")
        for item in (analysis.get("evidence") or [])
        if isinstance(item, Mapping)
    )


def _matched_markers(text: str, markers: Iterable[str]) -> list[str]:
    normalized_text = "".join(str(text or "").split())
    return [
        marker
        for marker in markers
        if "".join(marker.split()) in normalized_text
    ]


def _analysis_date_facts(analysis: Mapping[str, Any]) -> list[date]:
    return [
        parsed
        for item in (analysis.get("date_facts") or [])
        if isinstance(item, Mapping)
        if (parsed := _date(item.get("date"))) is not None
    ]


def _reference_dates(
    row: Mapping[str, Any],
    analysis: Mapping[str, Any],
) -> list[date]:
    values = [
        _date(row.get(field))
        for field in (
            "record_date",
            "pay_date",
            "share_arrival_date",
        )
    ]
    values.extend(_analysis_date_facts(analysis))
    strong_dates = sorted({value for value in values if value is not None})
    if strong_dates:
        return strong_dates
    announcement_date = _date(row.get("announcement_date"))
    return [announcement_date] if announcement_date else []


def _nearest_tdx_rows(
    rows: list[dict[str, Any]],
    reference_dates: list[date],
    *,
    limit: int = 3,
) -> list[dict[str, Any]]:
    if not reference_dates:
        return []
    ranked = []
    for row in rows:
        ex_date = _date(row.get("ex_date"))
        if ex_date is None:
            continue
        gap = min(abs((ex_date - anchor).days) for anchor in reference_dates)
        ranked.append({**row, "nearest_gap_days": gap})
    return sorted(
        ranked,
        key=lambda item: (
            int(item["nearest_gap_days"]),
            str(item.get("ex_date") or ""),
            int(item.get("id") or 0),
        ),
    )[: max(0, int(limit))]


def _tdx_economic_match(
    row: Mapping[str, Any],
    tdx: Mapping[str, Any],
    *,
    tolerance: float = 0.0001,
) -> bool:
    pairs = (
        (
            _per_ten(row.get("cash_dividend_per_share")),
            _number(tdx.get("fenhong")),
        ),
        (
            (
                (_per_ten(row.get("bonus_shares_per_share")) or 0)
                + (
                    _per_ten(
                        row.get("capitalization_shares_per_share")
                    )
                    or 0
                )
            ),
            _number(tdx.get("songzhuangu")),
        ),
        (
            _per_ten(row.get("rights_shares_per_share")),
            _number(tdx.get("peigu")),
        ),
        (
            _number(row.get("rights_price")),
            _number(tdx.get("peigujia")),
        ),
    )
    compared = False
    for left, right in pairs:
        left_value = left or 0.0
        right_value = right or 0.0
        if left_value or right_value:
            compared = True
        if abs(left_value - right_value) > max(0.0, tolerance):
            return False
    return compared


def _date_facts_text(analysis: Mapping[str, Any]) -> str:
    values = []
    for item in analysis.get("date_facts") or []:
        if not isinstance(item, Mapping):
            continue
        values.append(
            " | ".join(
                part
                for part in (
                    str(item.get("date") or ""),
                    str(item.get("date_type") or ""),
                    str(item.get("date_basis") or ""),
                )
                if part
            )
        )
    return _compact(values)


def _economic_terms_text(analysis: Mapping[str, Any]) -> str:
    terms = analysis.get("economic_terms")
    if not isinstance(terms, Mapping):
        return ""
    values = []
    for field, item in terms.items():
        if isinstance(item, Mapping) and item.get("value") is not None:
            values.append(
                f"{field}={item.get('value')} {item.get('unit') or ''}".strip()
            )
    return "; ".join(values)


def _review_reason_text(analysis: Mapping[str, Any]) -> str:
    review = analysis.get("_review_classification")
    if not isinstance(review, Mapping):
        return ""
    return _compact(
        [
            "原因代码: " + ", ".join(review.get("reason_codes") or []),
            "系统摘要: " + "；".join(review.get("operator_summary") or []),
        ]
    )


def _suggestion(
    *,
    state: str,
    event_kind: str,
    event_category: str,
    analysis: Mapping[str, Any],
    no_ex_right_markers: list[str],
    price_adjustment_markers: list[str],
    ex_date_markers: list[str],
) -> tuple[str, str, str]:
    if state == "evidence_unavailable":
        return (
            "E_需要补证据",
            "不可自动",
            "仅针对本事项重试公告发现；仍无证据时再决定是否采用外部日期旁证。",
        )
    if state == "document_rework":
        return (
            "F_需要修复文档",
            "不可自动",
            "只修复现有候选文档上下文并重新解析，不扩大公告扫描。",
        )
    if (
        event_kind == "送转配"
        and no_ex_right_markers
        and str(analysis.get("event_stage") or "")
        in {"implemented", "completed"}
    ):
        return (
            "A_官方明确不除权",
            "可按既定规则核准",
            "保留CNInfo事件和经济数字，factor_effect=none；事件日期采用官方登记、到账或上市事实，不使用TDX经济数字。",
        )
    if event_kind == "送转配" and price_adjustment_markers:
        return (
            "B_官方明确特殊价格调整",
            "需写入特殊因子依据",
            "保留CNInfo事件和经济数字；有效日及复权效果采用官方除权日或调整后的开盘参考价，不能套用普通全股东送转公式。",
        )
    if event_kind == "送转配" and event_category in SPECIAL_CATEGORIES:
        date_note = (
            "公告已明确除权日，但未提供特殊参考价或公式；"
            if ex_date_markers
            else ""
        )
        return (
            "C_特殊事项待定",
            "不可自动",
            "按非对称事项保留CNInfo数字；"
            f"{date_note}"
            "需判断复权效果及采用哪个官方实施日期，"
            "不能因TDX无记录直接推定factor_effect=none。",
        )
    if event_kind == "纯现金":
        return (
            "D_现金分派单独审核",
            "不可自动",
            "不能套用非对称送转规则；优先核准实际除息交易日和全股东现金口径。",
        )
    return (
        "G_其他冲突",
        "不可自动",
        "现有证据不足以按既定规则核准，需逐项确认事件身份、日期或条款。",
    )


def _load_blockers(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT s.instrument_id, i.name AS instrument_name, s.exchange,
               s.source_event_key, s.resolution_state, s.state_reason,
               s.next_action, s.candidate_count, s.latest_analysis_id,
               o.action_type, o.fiscal_period, o.announcement_date,
               o.record_date, o.ex_date, o.pay_date, o.share_arrival_date,
               o.cash_dividend_per_share, o.bonus_shares_per_share,
               o.capitalization_shares_per_share, o.rights_shares_per_share,
               o.rights_price, o.currency, o.description, o.event_status,
               o.quality_status, o.raw_payload_json,
               a.validation_status, a.error_code, a.result_json,
               a.gate_results_json
        FROM corporate_action_resolution_states s
        JOIN corporate_action_observations o
          ON o.instrument_id = s.instrument_id
         AND o.source_event_key = s.source_event_key
         AND o.source = 'cninfo'
         AND o.is_current = 1
        LEFT JOIN instruments i
          ON i.instrument_id = s.instrument_id
        LEFT JOIN corporate_action_llm_analyses a
          ON a.id = s.latest_analysis_id
        WHERE s.factor_blocking = 1
        ORDER BY s.resolution_state, s.instrument_id, s.source_event_key
        """
    ).fetchall()
    return [dict(row) for row in rows]


def _load_tdx(
    connection: sqlite3.Connection,
    instrument_ids: list[str],
) -> dict[str, list[dict[str, Any]]]:
    placeholders = ", ".join("?" for _ in instrument_ids)
    rows = connection.execute(
        f"""
        SELECT id, instrument_id, ex_date, fenhong, songzhuangu, peigu,
               peigujia, factor, validation_result
        FROM adjustment_factors_tdx
        WHERE instrument_id IN ({placeholders})
        ORDER BY instrument_id, ex_date, id
        """,
        instrument_ids,
    ).fetchall()
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        item = dict(row)
        grouped.setdefault(str(item["instrument_id"]), []).append(item)
    return grouped


def _load_evidence(
    connection: sqlite3.Connection,
    source_event_keys: list[str],
) -> list[dict[str, Any]]:
    placeholders = ", ".join("?" for _ in source_event_keys)
    rows = connection.execute(
        f"""
        SELECT instrument_id, source_event_key, evidence_source,
               resolution_status, announcement_id, announcement_title,
               announcement_time, effective_date, date_basis, evidence_url,
               raw_payload_json
        FROM corporate_action_effective_date_evidence
        WHERE source_event_key IN ({placeholders})
          AND observation_source = 'cninfo'
        ORDER BY source_event_key,
                 CASE resolution_status
                     WHEN 'resolved' THEN 0
                     WHEN 'candidate' THEN 1
                     ELSE 2
                 END,
                 announcement_time DESC,
                 id DESC
        """,
        source_event_keys,
    ).fetchall()
    return [dict(row) for row in rows]


def _build_report(
    blockers: list[dict[str, Any]],
    tdx_by_instrument: dict[str, list[dict[str, Any]]],
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, int]]:
    main_rows: list[dict[str, Any]] = []
    nearest_rows: list[dict[str, Any]] = []
    for row in blockers:
        analysis = _json_object(row.get("result_json"))
        raw = _json_object(row.get("raw_payload_json"))
        evidence_text = _evidence_text(analysis)
        no_ex_right = _matched_markers(
            evidence_text,
            NO_EX_RIGHT_MARKERS,
        )
        beneficiary_markers = _matched_markers(
            evidence_text,
            ASYMMETRIC_BENEFICIARY_MARKERS,
        )
        price_adjustment_markers = _matched_markers(
            evidence_text,
            PRICE_ADJUSTMENT_MARKERS,
        )
        ex_date_markers = _matched_markers(
            evidence_text,
            EX_DATE_MARKERS,
        )
        has_share_terms = any(
            (_number(row.get(field)) or 0) > 0
            for field in (
                "bonus_shares_per_share",
                "capitalization_shares_per_share",
                "rights_shares_per_share",
            )
        )
        has_cash = (_number(row.get("cash_dividend_per_share")) or 0) > 0
        event_kind = (
            "送转配"
            if has_share_terms
            else "纯现金"
            if has_cash
            else "其他"
        )
        event_category = str(raw.get("分红类型") or "").strip()
        if not event_category and has_share_terms:
            event_category = "特殊送转（原始分类缺失）"
        references = _reference_dates(row, analysis)
        nearest = _nearest_tdx_rows(
            tdx_by_instrument.get(str(row["instrument_id"]), []),
            references,
        )
        nearest_one = nearest[0] if nearest else {}
        nearest_gap = nearest_one.get("nearest_gap_days")
        nearest_economics_match = bool(
            nearest_one and _tdx_economic_match(row, nearest_one)
        )
        tdx_match_conclusion = (
            "无TDX记录"
            if not nearest
            else "邻近且经济一致"
            if int(nearest_gap) <= 7 and nearest_economics_match
            else "邻近但经济不一致"
            if int(nearest_gap) <= 7
            else "非邻近，仅供排查"
        )
        path, auto_approval, suggested_action = _suggestion(
            state=str(row["resolution_state"]),
            event_kind=event_kind,
            event_category=event_category,
            analysis=analysis,
            no_ex_right_markers=no_ex_right,
            price_adjustment_markers=price_adjustment_markers,
            ex_date_markers=ex_date_markers,
        )
        review = analysis.get("_review_classification")
        review = review if isinstance(review, Mapping) else {}
        main_rows.append(
            {
                "处理路径": path,
                "可否按既定规则直接核准": auto_approval,
                "建议动作": suggested_action,
                "事项键": row["source_event_key"],
                "证券代码": row["instrument_id"],
                "证券名称": row.get("instrument_name") or "",
                "市场": row["exchange"],
                "当前状态": row["resolution_state"],
                "系统原因代码": row["state_reason"],
                "系统下一动作": row["next_action"],
                "候选公告数": row["candidate_count"],
                "事件类别": event_category,
                "事项类型": event_kind,
                "原始事项说明": row.get("description") or "",
                "财年": row.get("fiscal_period") or "",
                "原始公告日": _date_text(row.get("announcement_date")),
                "原始登记日": _date_text(row.get("record_date")),
                "原始除权日": _date_text(row.get("ex_date")),
                "原始派息日": _date_text(row.get("pay_date")),
                "原始股份到账日": _date_text(row.get("share_arrival_date")),
                "CNInfo现金/10股": _per_ten(
                    row.get("cash_dividend_per_share")
                ),
                "CNInfo送股/10股": _per_ten(
                    row.get("bonus_shares_per_share")
                ),
                "CNInfo转增/10股": _per_ten(
                    row.get("capitalization_shares_per_share")
                ),
                "CNInfo配股/10股": _per_ten(
                    row.get("rights_shares_per_share")
                ),
                "CNInfo配股价": _number(row.get("rights_price")),
                "LLM分析状态": row.get("validation_status") or "无分析",
                "LLM事件阶段": analysis.get("event_stage") or "",
                "LLM事件类型": analysis.get("event_type") or "",
                "LLM置信度": analysis.get("confidence"),
                "LLM有效日": analysis.get("effective_date") or "",
                "LLM有效日类型": analysis.get("effective_date_type") or "",
                "LLM日期事实": _date_facts_text(analysis),
                "LLM经济条款": _economic_terms_text(analysis),
                "审核原因": _review_reason_text(analysis),
                "官方原文是否明确不除权": "是" if no_ex_right else "否",
                "不除权原文标记": "、".join(no_ex_right),
                "非对称受益原文标记": "、".join(beneficiary_markers),
                "特殊价格调整原文标记": "、".join(
                    price_adjustment_markers
                ),
                "明确除权日原文标记": "、".join(ex_date_markers),
                "关键官方原文": _compact(
                    [
                        item.get("exact_quote")
                        for item in (analysis.get("evidence") or [])
                        if isinstance(item, Mapping)
                        and (
                            _matched_markers(
                                str(item.get("exact_quote") or ""),
                                NO_EX_RIGHT_MARKERS,
                            )
                            or _matched_markers(
                                str(item.get("exact_quote") or ""),
                                ASYMMETRIC_BENEFICIARY_MARKERS,
                            )
                            or _matched_markers(
                                str(item.get("exact_quote") or ""),
                                PRICE_ADJUSTMENT_MARKERS,
                            )
                            or _matched_markers(
                                str(item.get("exact_quote") or ""),
                                EX_DATE_MARKERS,
                            )
                        )
                    ],
                    limit=1800,
                ),
                "最近TDX记录ID": nearest_one.get("id") or "",
                "最近TDX除权日": _date_text(nearest_one.get("ex_date")),
                "最近TDX距参考日/天": nearest_one.get(
                    "nearest_gap_days", ""
                ),
                "最近TDX现金/10股": nearest_one.get("fenhong"),
                "最近TDX送转/10股": nearest_one.get("songzhuangu"),
                "最近TDX配股/10股": nearest_one.get("peigu"),
                "最近TDX配股价": nearest_one.get("peigujia"),
                "TDX邻近匹配结论": (
                    tdx_match_conclusion
                ),
                "用户决定": "",
                "核准有效日": "",
                "factor_effect": "",
                "用户说明": "",
                "审核层级": review.get("review_tier") or "",
            }
        )
        for rank, tdx in enumerate(nearest, start=1):
            nearest_rows.append(
                {
                    "事项键": row["source_event_key"],
                    "证券代码": row["instrument_id"],
                    "证券名称": row.get("instrument_name") or "",
                    "参考日期": "、".join(item.isoformat() for item in references),
                    "距离排序": rank,
                    "TDX记录ID": tdx.get("id"),
                    "TDX除权日": _date_text(tdx.get("ex_date")),
                    "距最近参考日/天": tdx.get("nearest_gap_days"),
                    "TDX现金/10股": tdx.get("fenhong"),
                    "TDX送转/10股": tdx.get("songzhuangu"),
                    "TDX配股/10股": tdx.get("peigu"),
                    "TDX配股价": tdx.get("peigujia"),
                    "TDX因子": tdx.get("factor"),
                    "TDX校验状态": tdx.get("validation_result"),
                }
            )
    main = pd.DataFrame(main_rows).sort_values(
        ["处理路径", "市场", "证券代码", "事项键"],
        kind="stable",
    )
    nearest_df = pd.DataFrame(nearest_rows)
    path_counts = Counter(str(value) for value in main["处理路径"])
    return main, nearest_df, dict(sorted(path_counts.items()))


def _evidence_frame(
    evidence_rows: list[dict[str, Any]],
) -> pd.DataFrame:
    items = []
    for row in evidence_rows:
        if str(row.get("resolution_status") or "") not in {
            "candidate",
            "resolved",
        }:
            continue
        payload = _json_object(row.get("raw_payload_json"))
        classification = payload.get("title_classification")
        classification = (
            classification if isinstance(classification, Mapping) else {}
        )
        items.append(
            {
                "事项键": row["source_event_key"],
                "证券代码": row["instrument_id"],
                "证据来源": row["evidence_source"],
                "证据状态": row["resolution_status"],
                "公告ID": row.get("announcement_id") or "",
                "公告标题": row.get("announcement_title") or "",
                "公告日期": _date_text(row.get("announcement_time")),
                "证据有效日": _date_text(row.get("effective_date")),
                "日期依据": row.get("date_basis") or "",
                "标题分类角色": classification.get(
                    "announcement_role", ""
                ),
                "标题分类原因": classification.get("reason", ""),
                "证据链接": row.get("evidence_url") or "",
            }
        )
    return pd.DataFrame(items)


def _style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    decision_fill = PatternFill("solid", fgColor="FFF2CC")
    danger_fill = PatternFill("solid", fgColor="FCE4D6")
    for worksheet in writer.book.worksheets:
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        worksheet.row_dimensions[1].height = 36
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
        for column_index, column in enumerate(
            worksheet.iter_cols(min_row=1),
            start=1,
        ):
            values = [str(cell.value or "") for cell in column[:100]]
            width = min(
                55,
                max(10, max((len(value) for value in values), default=10) + 2),
            )
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width
        headers = {
            str(cell.value): cell.column
            for cell in worksheet[1]
            if cell.value
        }
        for header in ("用户决定", "核准有效日", "factor_effect", "用户说明"):
            column_index = headers.get(header)
            if not column_index:
                continue
            for column in worksheet.iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
            ):
                for cell in column:
                    cell.fill = decision_fill
        path_column = headers.get("处理路径")
        if path_column:
            for row_index in range(2, worksheet.max_row + 1):
                path_value = str(
                    worksheet.cell(row=row_index, column=path_column).value
                    or ""
                )
                worksheet.cell(
                    row=row_index,
                    column=path_column,
                ).fill = (
                    section_fill
                    if path_value.startswith("A_")
                    else danger_fill
                )


def export_report(database_path: Path, output_path: Path) -> dict[str, Any]:
    uri = f"file:{database_path.resolve()}?mode=ro"
    with sqlite3.connect(uri, uri=True) as connection:
        connection.row_factory = sqlite3.Row
        blockers = _load_blockers(connection)
        if not blockers:
            raise RuntimeError("no factor-blocking CNInfo events found")
        instrument_ids = sorted(
            {str(row["instrument_id"]) for row in blockers}
        )
        event_keys = sorted(
            {str(row["source_event_key"]) for row in blockers}
        )
        tdx_by_instrument = _load_tdx(connection, instrument_ids)
        evidence_rows = _load_evidence(connection, event_keys)

    evidence_status_counts = Counter(
        str(row.get("resolution_status") or "")
        for row in evidence_rows
    )
    main, nearest, path_counts = _build_report(
        blockers,
        tdx_by_instrument,
    )
    evidence = _evidence_frame(evidence_rows)
    state_counts = dict(
        sorted(Counter(main["当前状态"].astype(str)).items())
    )
    summary = pd.DataFrame(
        [
            ["生成时间", datetime.now().isoformat(timespec="seconds")],
            ["数据库", str(database_path)],
            ["阻塞事项总数", len(main)],
            ["涉及证券数", main["证券代码"].nunique()],
            ["官方明确不除权候选", path_counts.get("A_官方明确不除权", 0)],
            [
                "TDX邻近且经济一致",
                int(
                    (
                        main["TDX邻近匹配结论"]
                        == "邻近且经济一致"
                    ).sum()
                ),
            ],
            ["候选公告明细", len(evidence)],
            [
                "已过滤的拒绝公告",
                int(evidence_status_counts.get("rejected", 0)),
            ],
            [
                "结论",
                (
                    "不得直接计算最终因子。先核准阻塞事项；"
                    "黄色列供用户批示。A类已有官方明确不除权原文，"
                    "可按factor_effect=none路径审核。"
                ),
            ],
            [
                "数据访问",
                "只读本地SQLite；未下载公告、未运行OCR或LLM、未写审核记录。",
            ],
        ],
        columns=["项目", "内容"],
    )
    path_summary = pd.DataFrame(
        [
            {
                "处理路径": path,
                "数量": count,
                "证券代码": "、".join(
                    main.loc[main["处理路径"] == path, "证券代码"]
                    .drop_duplicates()
                    .astype(str)
                ),
            }
            for path, count in path_counts.items()
        ]
    )
    state_summary = pd.DataFrame(
        [
            {"当前状态": state, "数量": count}
            for state, count in state_counts.items()
        ]
    )
    fields = pd.DataFrame(
        [
            ["处理路径", "系统建议的处置分组，不等于最终用户决定。"],
            [
                "A_官方明确不除权",
                "现有官方原文明确表示不除权，候选factor_effect=none。",
            ],
            [
                "B_官方明确特殊价格调整",
                "官方明确开盘参考价调整或特殊计算公式，需保存特殊因子依据。",
            ],
            [
                "明确除权日原文标记",
                "只证明公告明确了除权日，不等于已获得特殊参考价或公式。",
            ],
            [
                "最近TDX距参考日/天",
                "最近TDX除权日与CNInfo/LLM已知日期的最小自然日差。",
            ],
            [
                "TDX邻近匹配结论",
                "7日以内仅标记邻近，仍不代表经济数字可用于CNInfo。",
            ],
            [
                "关键官方原文",
                "只摘录命中不除权或非对称受益关键词的已有LLM证据。",
            ],
            [
                "factor_effect",
                "用户批示字段：normal或none；空白表示尚未决定。",
            ],
        ],
        columns=["字段", "含义"],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="00_总览", index=False)
        path_summary.to_excel(
            writer,
            sheet_name="00_总览",
            index=False,
            startrow=len(summary) + 3,
        )
        state_summary.to_excel(
            writer,
            sheet_name="00_总览",
            index=False,
            startrow=len(summary) + len(path_summary) + 7,
        )
        main.to_excel(writer, sheet_name="01_63条逐项", index=False)
        nearest.to_excel(writer, sheet_name="02_TDX最近记录", index=False)
        evidence.to_excel(writer, sheet_name="03_公告证据清单", index=False)
        fields.to_excel(writer, sheet_name="04_字段说明", index=False)
        _style_workbook(writer)

    return {
        "status": "success",
        "output_path": str(output_path),
        "blocker_count": len(main),
        "instrument_count": int(main["证券代码"].nunique()),
        "path_counts": path_counts,
        "state_counts": state_counts,
        "evidence_rows": len(evidence),
        "rejected_evidence_rows_excluded": int(
            evidence_status_counts.get("rejected", 0)
        ),
        "tdx_comparison_rows": len(nearest),
        "network_access": False,
        "llm_invocations": 0,
        "database_writes": 0,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--database",
        type=Path,
        default=Path("data/quotes.db"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(
            "data/reports/"
            "cninfo_corporate_action_blockers_review_20260727.xlsx"
        ),
    )
    args = parser.parse_args()
    result = export_report(args.database, args.output)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
