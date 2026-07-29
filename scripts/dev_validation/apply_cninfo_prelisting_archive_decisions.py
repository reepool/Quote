#!/usr/bin/env python3
"""Apply fixed CNInfo archive dispositions and export ten manual-review gaps."""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sqlite3
import sys
from collections import Counter
from datetime import date
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable, Mapping

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from data_sources.cninfo_factor_governance import (  # noqa: E402
    match_cninfo_archive_tdx_date,
)


DEFAULT_DATABASE = ROOT_DIR / "data/quotes.db"
DEFAULT_REPORT = (
    ROOT_DIR
    / "data/reports/cninfo_post_listing_archive_gaps_review_20260729.xlsx"
)
REVIEWER = "operator_cninfo_archive_cleanup_20260729"
EXPECTED_DECISION_EVENT_KEYS_HASH = "2e156d20e4036950fc9c2a114a8cf9896e2034b194c6f53479f526834338efe4"
EXPECTED_DECISION_OBSERVATION_MANIFEST_HASH = "609f5cb40ba9c14ff679781642e7897779fcb13681bd50b12971a320ff9fa958"
EXPECTED_DECISION_PAYLOAD_HASH = "06902714ea32e4401a8312a2796d1dcafb3e8d6ec5c620f99f47881e6f527bf3"
EXPECTED_REVIEW_EVENT_KEYS_HASH = "41c12107c50ee0a6e9896744ace7411e005ef7efe0648cb5057e44e91bab1c31"
EXPECTED_REVIEW_OBSERVATION_MANIFEST_HASH = "15198903e896e529c920dac67c2c7a5d16d9e06416c777b8273986d491816ce3"

ECONOMIC_FIELDS = (
    "cash_dividend_per_share",
    "bonus_shares_per_share",
    "capitalization_shares_per_share",
    "rights_shares_per_share",
    "rights_price",
)
PRODUCTION_FACTOR_TABLES = (
    "adjustment_factor_observations",
    "adjustment_factors",
    "adjustment_factors_canonical",
    "adjustment_factor_instrument_status",
)

# These are operator-confirmed exceptions, not a generic announcement-date rule.
# Each tuple is instrument, source event key, current CNInfo row hash, basis.
PRE_LISTING_SPECS: tuple[tuple[str, str, str, str], ...] = (
    ("000420.SZ", "4a3d052e63e70e2f5b1381e29bb6818deb9602403156fca9779b2205fa6172c6", "a669dd8ca85bdc09f1db722a45630f38ff6264988776b6977cc4f3407a916f52", "operator_confirmed_announcement_only"),
    ("000421.SZ", "ec1aabda0ba73286176770ffc13c6d8d44929120ff723b309a429ad663c7ae90", "9f6590629bb747f0e4c62c977fbb9955c055b68146c7d7f85b1ff0301e165f42", "operator_confirmed_announcement_only"),
    ("000429.SZ", "2d1633e87e1c264ddc42f75c59155cc3d0116b346fa85a3a1f8b20cd6e1aef0a", "81b0e5421581b80ebd3f38da637520483d4937c62b830cf2364e6afcbfb14024", "operator_confirmed_announcement_only"),
    ("000429.SZ", "5dae9e0dcdeb9e2419cbd72c5a84dd48cf84e4c989f606f674ed67de256e1155", "1b131c8f6d9a4ee66b36812f4fcb5556faef44f790736f15e165d9cd40a61512", "operator_confirmed_announcement_only"),
    ("000488.SZ", "1bd899821dc72ca417e291c2bb926eecbf976115d52ee094e7328b59cf114389", "5f38813d604d749511111773a4041ff3bb2fa48bdf8c1c9190040b491f76198e", "operator_confirmed_announcement_only"),
    ("000488.SZ", "3e7f00ae45a839104182abde52a9d970b6a8f882fed708e53cb3b1bb687b6415", "c8515d93f1b5fef0e8afe4bdadab83d5a5d491494621892126540dba187a5b18", "operator_confirmed_announcement_only"),
    ("000600.SZ", "3fa325d201025f00bb22fd45dfeba4423ce719d9e43adc4c8f580bd2d7acd3a7", "d47be6bb65144a72c1fe039c76f52978e98e5017bb6be2e8938c0c44032a96fc", "operator_confirmed_announcement_only"),
    ("000609.SZ", "f31585daaaa2ac2789bd3f4c6ff0d2808f11b52472ee52bb644e48b0a3d8557b", "39ba9e550dd36943a0c169ceedecd67e064bace7aaac549cafbd8d4386872e7b", "operator_confirmed_announcement_only"),
    ("000612.SZ", "ffc6e2c2fd4360069e0166fc12cf07ac862e496c9c9d347a4a4ea603b81cd237", "1ec7f7fc2359d08c00c99bcead57eede2748a645dfcc33bdc4e9fb9bd1be766c", "operator_confirmed_announcement_only"),
    ("000630.SZ", "c836810fd61f5ec8b8d190a69ce4074116fa4de843f6ed2ab0a67a271efe3b92", "05eeb793903592065a9da139d1b1af4e28e8c3a6939bb725a96d35d9b1f54d20", "operator_confirmed_announcement_only"),
    ("000631.SZ", "70752badc32b844aa1a1e56758e7f39ba28a87eb8936d0bf3c07720f8caa5fc6", "241869f03ff85773f40421fb248a8d2e43d18d1b355f2f6e94a5ee32e6238a1a", "operator_confirmed_announcement_only"),
    ("000656.SZ", "bf7cf7e8da9a73717ed845ed30307c16937f3ce01aab93f7306738ac512e641f", "2e938222a350c5e5020b723abcc3a1fa3d60b9ffca3ebcc00b9ad16f7f580c30", "operator_confirmed_announcement_only"),
    ("000669.SZ", "667627481b93e569a122bcefa3e848948caeaddb564fc785b3093e07c7a59a7b", "bba7d6c8ae3f9287b3557e52c2faecd6716d5aa026fcccb7d4183fdef7c81e77", "operator_confirmed_announcement_only"),
    ("000725.SZ", "9b788dc8ba72581a1d36df8cec2d5e985ecfca6f443e13dd2818dcbab624b578", "81da99663b7e0121ccd8859a0577e06368b255ed66cca3e90505d31f992658f1", "operator_confirmed_announcement_only"),
    ("000726.SZ", "130acfc4d0b52440433b6be90adfd4157e9ffa4d51eab908d57c3557f12929e3", "f9eed81f594ba5b817fcddc94326b13b4792f545f44f0511aaa0f5c921282734", "operator_confirmed_announcement_only"),
    ("000726.SZ", "1db85fe202faed3481af6c2eccc4cb14c15fedb9b9b1470a73ca79560febe261", "4a112fed3b170f87e264a54da5890bb4d3555ea347ee121db0a3fbb16b2e5483", "operator_confirmed_announcement_only"),
    ("000726.SZ", "f4e1d9132b7421799d04de7c45269df27d969309561d5ebcea573f8efac8ce67", "3d2e2020ac4f3246d8524551117c20c1c36c40c77063229743fdd51dada23251", "operator_confirmed_announcement_only"),
    ("000869.SZ", "15ea8cc9dd5759a9ba30f1e6e3a22ff957ba4a9ee0d1b1fe547b161a6b200c60", "c5fc2aead65145d2524ed435c5c70701d202c568884f898f26f93da03f585162", "operator_confirmed_announcement_only"),
    ("600221.SH", "b4647b52a366c55b963b09bb1cce0f739b5d1cfe0927e8532ba79d7f2148b509", "8c3afeaacb9bd8f2fb91d8ceb9d55f4c76c9b0021bf29f2042724c355fbf1102", "operator_confirmed_announcement_only"),
    ("600272.SH", "20e4bb972385d6e927dcc37ff443f8922f11d76a2a5bf32d296e607d8145f022", "6ea623618ebe4f184d5609b6e854c94dd1b6f9392182eb3b6669f8166e0d9a8d", "operator_confirmed_announcement_only"),
    ("600272.SH", "3e1a5352a2c52cc35991addd520438e0dea708d3d7331d50660ffce191e03719", "0248a92fc40ea36d7dd90ce58fa66604ebb2835657c579ca50515d9426860278", "operator_confirmed_announcement_only"),
    ("600295.SH", "037a9b0be7cc1cda764e4ed7dbe245578f6552d51c2dca9bcbc52ab579eb6502", "4fa19bd37e2bab92854bd90f58ce8a46363040e3c23ffcc04f5cbd9e37722676", "operator_confirmed_announcement_only"),
    ("600295.SH", "6ecf0c6bb0ea4e9631799a723d371ce4d5361ec98646771bda07070762ef4a39", "b3a73bd8346f4a40931613096e7ec5e33e8c6519f5c9a54df465ee5b431d468f", "operator_confirmed_announcement_only"),
    ("600295.SH", "abb9a6c80f9fbbf9481598a1ac3f3de70940ced2650195e95d18e7a6970c13a9", "a31e9cd0c66fb24d2500b31d18fda7b3eeee59cb284428e217ddd0d061458beb", "operator_confirmed_announcement_only"),
    ("600295.SH", "d7b29586cf94631cd99ebb98f734c3026d366e2dbb95a1dba3f1ad4fb550dee4", "65f95689892294a5acbc7fc30e3c14ee6b5571d6c66cc93f6a0ba19f333cefec", "operator_confirmed_announcement_only"),
    ("600320.SH", "0a326eac2eb1d4c8a17d24a83e5b4fd1aea2f8d73df7dbcff08f8cf6aa1c4d97", "5f1f5baa5b95bd07ca6c090837f4626055b4432e35973724c009f45e5fbeb104", "operator_confirmed_announcement_only"),
    ("600320.SH", "95e4691682af412d0fd36b606715181d968e34410b86d2c2c9ec66d00a0e3481", "79862b1cd90d901f319f71096615881bf6e11620f829de1316c9ab9df8abc0c8", "operator_confirmed_announcement_only"),
    ("600320.SH", "c3717331ef62fd277e19c854784a93e2b32618ed17ee4ba4607872239db9e790", "a9ba700f05607210a4f65941f9527b82637c06299bdd06be656e072abd300e05", "operator_confirmed_announcement_only"),
    ("600704.SH", "c351003cc66ee60777e1dfceaed55e7149938a28e9d22dd9d69077fdf83436b1", "18f4ba07aa3ee328bf8f1d73cd0aee1ea06263f9d2e4dd0ea77442e5552959a9", "operator_confirmed_announcement_only"),
    ("600712.SH", "e66728e3b48179ae70b001eea84176fbec4fd771ff07e360f9efab0bfaab6a66", "6365ad04c896a84a6ddcd24d6753b1c0cbb07d140fe77075e3234b6334a63326", "operator_confirmed_announcement_only"),
    ("600715.SH", "038df3dc5edabca45736808997a9374a636edbe94cb383a953a4979f36517c45", "ef06308634e0d3e1a6d8880a8db7a2084f2e1456d2f729392812f0f313ae4015", "operator_confirmed_announcement_only"),
    ("600722.SH", "3f74425ca759e5cea60e9b3c7287178c06ab308add883c4d8eb32c421ac19fbf", "a6fe69cce289bb9913a22be1200d244ca11650c417ef7021b0c0d0b94c4fa821", "operator_confirmed_announcement_only"),
    ("600726.SH", "6759f5abf4ee290d7cf5a45c15c4cb0bafe9c30e7174a7d92945d6dd24773a3c", "6588247627fd18a51ec2726e6df0418ae1cc8f596f4a58a831b8a5fa63596d63", "operator_confirmed_announcement_only"),
    ("600731.SH", "5c8e667b345e2c81ebf91304392471fa814d524578c36a0c3536692c280460a4", "c9ad99f16c1fa7e199b344542240581b9a4eae30fabdbc260953dfac2215fc1c", "operator_confirmed_announcement_only"),
    ("600734.SH", "3dff144df7f92496bd54549a36eaa31087e9b0bd372cf4aecc5d2e8d5b660e74", "7def01eb2c437db83a4b42dc3d431f4760f9842a76b151a8afb4863ea02fe203", "operator_confirmed_announcement_only"),
    ("600737.SH", "f37852815570a264e5f2a1ff834bd1a925b5ad1548842b1ebc712417fc41f747", "ff1eb28be8c600294af6bca80e926c9242912356644d5ca48f4b04dc268a98d1", "operator_confirmed_announcement_only"),
    ("600739.SH", "9c84fa013eb5587ad05c4dc960e0cb030d19e3f4f147e188340390f90f178ffd", "7e2701f00e404f4baefe6ffc6d638f271edd1b0e78a9bd68074436e8443f261e", "operator_confirmed_announcement_only"),
    ("600756.SH", "647a041f3a9c2af6c8864fdc7dbf493d8694237aab410b80be8703748e379345", "6dcb87da48e2cdd4f69453572fe0fc7b3e554cf9b074e74e43d1220f5aff55f8", "operator_confirmed_announcement_only"),
    ("600760.SH", "3da6b1675aa1385bdffcd9a071142465b776a637fdb7b4ae57df412b9693d3a3", "c9403578ce14e769bad2e9b6c0022986c47bc654094f16d105fec1796e7d4129", "operator_confirmed_announcement_only"),
    ("600764.SH", "d056581e420a06aeae0f7c52427a903c23858c745c6e637307954b9e7d1cb751", "3b2723a0324c40fc3923f0eaa9b2e010f6188800c7f3a6bef4302e7fb48084fe", "operator_confirmed_announcement_only"),
    ("600769.SH", "64df5e0aa4902c0eb7540cfacd104b96c5870234a0beb0d50481682a901cd5aa", "6b0394405bd12cface980dd8df86bc1d95faeea1d2977296877edb8425f50bcc", "operator_confirmed_announcement_only"),
    ("600778.SH", "8be6b11e258cdf6c03f5a3882960e99bee2e39e34fa1e999483c25e44b895b40", "0c46e84363dc900a95a23a92daea93a84d45fad483bc21bc087a314be1b82853", "operator_confirmed_announcement_only"),
    ("000725.SZ", "d58076b14d185f8be108c609c33369af98f63b3a393bd757b9ae06d69b55f852", "57a3a18889187513b66435efba4af4a704ba98bf5c37323cc2c3716cc92b9865", "operational_dates_before_listing"),
    ("600272.SH", "da08b3c281050826f627d9589522624d6612559ad3f45c0aa5991be6d2b2e96f", "70e47c85ef41f9054c3aace70c79c0faca4ca015176ebcab13a6da49c87a85c9", "operational_dates_before_listing"),
    ("600295.SH", "1035a927c3e6f4910c00b72974d7ac9428626fa88a3155b27fecc9e13e09541d", "ccdabc0a2f358018d9424b088cbfdf5ba5fef6fa7f9cc48cd567d7199fe2146d", "operational_dates_before_listing"),
)

NON_EFFECTIVE_SPECS: tuple[tuple[str, str, str], ...] = (
    ("000408.SZ", "60a7a26ee623a238e08c261d87e2de0b5e9f9c06dc6779f24c7bbc345ab8b81f", "f349d95df632cf66f9f18ec13eb8bfe7fa588595c8698334642af4ba9289c901"),
    ("000425.SZ", "9ed0ab91435b9106cddcc0c032316660251c2687029454c3853ecab622709fc5", "675421fa11f009f9fb905d9cbcd6578d7a0133f2b4664a0df083e43e5b0278c3"),
    ("000608.SZ", "988079929e9a224aa0a2dd8a0fb45a236cfc3cddbe0b24b3690ff37b11017aea", "9e6fdf135a7221df8e18e052a5b63efcdd91b7071029d4f8e7024781912a1d3a"),
    ("000661.SZ", "9136f0a77bb79a1b16d870adf76e3f91aa41cfb7e6c82d909b5cf263a819e259", "130ffce9c9ebfb87a9b4e3e3311e2df9649a894b076b29456eb35d10ae6262ba"),
    ("000665.SZ", "4343afb6a0ff387f09c56ba91f49d40aee26dcdb035535058bfbc08e1d2e10a9", "94b5a8834906cec838dbf689c49018179c38435d5ca5060e4e7673141b357e85"),
    ("000668.SZ", "7357a1b700ce30601d7375d7e5652f9ba079f237b2e92c577ec4f6d524929032", "29a0d30c28fc6c786b5700263979c5c9d31f3d2134124ee5f646861d9505b60f"),
    ("600702.SH", "d7728f9aa1d2ddb21a13745972aaa73d0b0923e03a2697cb5cec6ee8e2edd722", "a3a5f19ba1280f0cdac2188da02e15b30bc76c5a8a2c081964b883112ec88ef6"),
    ("600703.SH", "622d245708672758a4c94d96bcb48fcc2340bebee136d8d0aaa30d03e8157516", "fe307ad7e0ad491edde929bfbb755e20e9d5360e9212484d0928fa54d40fc4f5"),
    ("600710.SH", "5350790826356fc1d026ba7fe2a85ff68c8c4505224517629aaab1a2b98836ab", "644b5cd63bd43929a4bd6b86d14cadabacc366bb36d7f6ab14d94db7136561dc"),
    ("600716.SH", "c408b625db10036b7c4949d9c5d633662aea4f6943847a31a4fdaccd8fbc6e24", "a297e08535d74a532faed81f21e1472641b2354adef6f6cf0480cb9fb79143f2"),
    ("600719.SH", "e498257effcc6b51193df8629dc337ba454a3702717501e2e537c5b3757c3b02", "f3d494adfb8e03a3ed1967de8bea999218c0d864367becd25bebc23767aea266"),
    ("600720.SH", "4284c672495f442a633d4d4ddda30d605952d869d8cc440d2fa14397e59f4c94", "99e4463c8d606f7ae3128dcfa9bd8ef0313811ec7d4dd7653c6ac34995c4df3f"),
    ("600721.SH", "ed04355da66fe1a61f339f36ab994b04430c27867072247d19780ef79e8905aa", "f99364b433fb1e0a17295cc6f671a5b1e798db18bcca97bed10e1c045c99ed21"),
    ("600728.SH", "366002ff52c86cb7bf35613dc3270bc52086ab4c041b6818ea4f53bd434f0ff8", "c90377352e5fa40e0f3bd17112f39303f4ff13a4e5de021fb37cf022abc03fa9"),
    ("600773.SH", "0ba6a2de6cdc8de73b1dda96cc01710c2982291d64a374e9f3530731a6c8958a", "fa72e32418dd3526f92155b57cb49c1b708df666a7b48ace1fd13df866632257"),
    ("600774.SH", "8c5eb7b981b5ffb7dc8caea2c0520ecf8d08e49179c264751fe8c44bd8d1d2bf", "72a81ea39b706588d172b2229c5676dcae0f8adce2b8e0f37acd339f7f709949"),
    ("600777.SH", "ab96428c211a396bd6be4daa92ecd083b3e4ae223ff640bad7992ccd6bcd5718", "1e473b4158c5badc17b5aebb8f66635b81efdbfa0d98bdd6dd3767c255da448a"),
)

REVIEW_SPECS: tuple[tuple[str, str, str], ...] = (
    ("000055.SZ", "c4477a3b305fc9eba9ce2a734de696881581484c7e3df0242b04a3dd3648db39", "fcbd65bcda14550f2a107dcc677c0c400f2c64d05140e98f30807fe57a9716b0"),
    ("000415.SZ", "890605e30b27ea66f1cc4152223ee765ac35802b35390343516894d02b156a27", "9c7b299ff6b582f6cd961246606df7be3e21b37830086d6823ea2d8e72146ef3"),
    ("000558.SZ", "b41e6d53d0666ce892d9daf7b957bbb62bf719ccc81c1d1f0c5e07c7d31814e8", "b01b633bea8fc2a13a7dccb7e6a7fe5cbd4034962806397fb9b3086925c2bd90"),
    ("000592.SZ", "514e7675bc7e56a98620976e60742833c248599ce03264a07d1e56f26b6e76ec", "85c5c612a320e5ee2d3be1acd51863f20dc31371bcae1adade1bfcaa08819211"),
    ("000625.SZ", "3e567fb54edfa7e3d37756a9ea09514582d4e4e6c94755a29bc0ccca0628bc83", "c6f44938fa0c8856eb8a3ff59ac809109980d079da76bed0b8f81f3e6d82bb57"),
    ("000793.SZ", "3572c075ae770e5c99fd8832477be8a6bcbe761f3045cbcf61b97b52d65d136b", "b3e26ebb77aae89d0717312b3b1d66ff85a9094b18aa0ddde41e0e6b3e928756"),
    ("600236.SH", "c3863809107219158c58c204aa54630d277ca6a55d924b53642cf28156ec25f4", "9d33a7cc4ac8c9f16c816d468c001c0f9bb3b75832409eb604c7602d07f7158a"),
    ("600611.SH", "486626f6f9e8eccc97b57beba97b49e4c36065fe83d0ae908662573265b32138", "0724b2c9314d21706be6a5c9f280a0245f147bd4a5332bfc1231fdeb0134c2e9"),
    ("600738.SH", "68da85219c4104aa3e6fd0946aecd8f8fbeca252c9679b4f839a444102f2aa5e", "b8aeef5e6fd5fcd9e6a6be76d0244ad2bf043b92bfa11c6cf8f67a590ad67285"),
    ("600887.SH", "806b3bcb247a8b9f2b8de9232c213bb4a76c10643de92281047e9a08da21c774", "72a7c84dc39cf1cb74f148317da39e6529c7df2af77b841115e1d3ea7890a6f2"),
)


def _canonical_hash(value: Any) -> str:
    return sha256(json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")).hexdigest()


def _hash_lines(values: Iterable[str]) -> str:
    return sha256("\n".join(sorted(values)).encode("utf-8")).hexdigest()


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _date_text(value: Any) -> str:
    parsed = _date(value)
    return parsed.isoformat() if parsed else ""


def _per_ten(value: Any) -> float | None:
    if value in (None, ""):
        return None
    return float(value) * 10.0


def _connect_read_only(path: Path) -> sqlite3.Connection:
    resolved = path.expanduser().resolve()
    if not resolved.is_file():
        raise FileNotFoundError(f"SQLite database does not exist: {resolved}")
    connection = sqlite3.connect(f"{resolved.as_uri()}?mode=ro", uri=True)
    connection.row_factory = sqlite3.Row
    return connection


def _single_current_observation(
    connection: sqlite3.Connection,
    instrument_id: str,
    source_event_key: str,
) -> dict[str, Any]:
    rows = connection.execute(
        """
        SELECT o.*, i.name AS instrument_name, i.listed_date,
               s.resolution_state, s.factor_blocking
        FROM corporate_action_observations AS o
        JOIN instruments AS i
          ON i.instrument_id=o.instrument_id
        LEFT JOIN corporate_action_resolution_states AS s
          ON s.instrument_id=o.instrument_id
         AND s.source_event_key=o.source_event_key
        WHERE o.instrument_id=? AND o.source_event_key=?
          AND o.source='cninfo' AND o.is_current=1
        """,
        (instrument_id, source_event_key),
    ).fetchall()
    if len(rows) != 1:
        raise RuntimeError(
            f"current CNInfo observation missing or ambiguous: "
            f"{source_event_key}"
        )
    return dict(rows[0])


def _load_tdx(
    connection: sqlite3.Connection,
    instrument_ids: Iterable[str],
) -> dict[str, list[dict[str, Any]]]:
    normalized = sorted(set(instrument_ids))
    placeholders = ",".join("?" for _ in normalized)
    rows = connection.execute(
        f"""
        SELECT id, instrument_id, ex_date, factor, cumulative_factor,
               validation_result, pre_close, fenhong, songzhuangu,
               peigu, peigujia
        FROM adjustment_factors_tdx
        WHERE instrument_id IN ({placeholders})
        ORDER BY instrument_id, ex_date, id
        """,
        normalized,
    ).fetchall()
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        item = dict(row)
        grouped.setdefault(str(item["instrument_id"]), []).append(item)
    return grouped


def _validate_hash(actual: str, expected: str, label: str) -> None:
    if expected and actual != expected:
        raise RuntimeError(f"{label} drifted: {actual}")


def build_decisions(
    connection: sqlite3.Connection,
) -> list[dict[str, Any]]:
    specs = [
        {
            "instrument_id": instrument_id,
            "source_event_key": event_key,
            "expected_row_hash": row_hash,
            "terminal_reason": "pre_listing",
            "basis": basis,
        }
        for instrument_id, event_key, row_hash, basis in PRE_LISTING_SPECS
    ]
    specs.extend({
        "instrument_id": instrument_id,
        "source_event_key": event_key,
        "expected_row_hash": row_hash,
        "terminal_reason": "non_effective",
        "basis": "no_positive_current_economic_term",
    } for instrument_id, event_key, row_hash in NON_EFFECTIVE_SPECS)
    if len(specs) != 62:
        raise RuntimeError("fixed terminal-decision manifest must contain 62 events")
    if len({item["source_event_key"] for item in specs}) != 62:
        raise RuntimeError("fixed terminal-decision event keys must be unique")
    tdx_by_instrument = _load_tdx(
        connection,
        [item["instrument_id"] for item in specs],
    )

    rows = []
    for spec in specs:
        observation = _single_current_observation(
            connection,
            spec["instrument_id"],
            spec["source_event_key"],
        )
        if observation["row_hash"] != spec["expected_row_hash"]:
            raise RuntimeError(
                "CNInfo observation row hash drifted: "
                + spec["source_event_key"]
            )
        current_state = str(
            observation.get("resolution_state") or ""
        ).strip()
        if current_state not in {
            "official_archive_unavailable",
            spec["terminal_reason"],
        }:
            raise RuntimeError(
                "terminal decision would overwrite a newer state: "
                f"{spec['source_event_key']}={current_state}"
            )
        listed_date = _date(observation.get("listed_date"))
        announcement_date = _date(observation.get("announcement_date"))
        operational_dates = [
            parsed
            for field in ("record_date", "pay_date", "share_arrival_date")
            if (parsed := _date(observation.get(field))) is not None
        ]
        if spec["terminal_reason"] == "pre_listing":
            if listed_date is None:
                raise RuntimeError(
                    "pre-listing decision lacks listing date: "
                    + spec["source_event_key"]
                )
            if spec["basis"] == "operator_confirmed_announcement_only":
                if operational_dates or not (
                    announcement_date and announcement_date < listed_date
                ):
                    raise RuntimeError(
                        "announcement-only pre-listing evidence drifted: "
                        + spec["source_event_key"]
                    )
                tdx_match = match_cninfo_archive_tdx_date(
                    observation,
                    tdx_by_instrument.get(spec["instrument_id"], []),
                    field_tolerance=0.0001,
                )
                if tdx_match.get("matched"):
                    raise RuntimeError(
                        "announcement-only pre-listing event gained a "
                        "unique TDX market match: "
                        + spec["source_event_key"]
                    )
            elif not operational_dates or max(operational_dates) >= listed_date:
                raise RuntimeError(
                    "operational pre-listing evidence drifted: "
                    + spec["source_event_key"]
                )
            notes = (
                "用户确认该固定事项属于上市前事项，不参与上市后二级市场"
                "复权；不以公告日伪造实施日。"
            )
        else:
            if any(
                float(observation.get(field) or 0) > 0
                for field in ECONOMIC_FIELDS[:-1]
            ):
                raise RuntimeError(
                    "non-effective observation gained a positive term: "
                    + spec["source_event_key"]
                )
            notes = (
                "历史记录仅说明利润滚存、归老股东或发起人等情况，不是"
                "当前上市市场已实施的分派送转事项。"
            )
        rows.append({
            **spec,
            "reviewer": REVIEWER,
            "notes": notes,
            "operator_attestation": {
                "basis": spec["basis"],
                "listed_date": _date_text(observation.get("listed_date")),
                "announcement_date": _date_text(
                    observation.get("announcement_date")
                ),
                "operational_dates": {
                    field: _date_text(observation.get(field))
                    for field in (
                        "record_date",
                        "pay_date",
                        "share_arrival_date",
                    )
                    if _date(observation.get(field)) is not None
                },
                "no_fabricated_effective_date": True,
                "network_access": False,
                "llm_invocations": 0,
            },
        })
    rows.sort(key=lambda item: (
        item["terminal_reason"],
        item["instrument_id"],
        item["source_event_key"],
    ))
    event_keys = {item["source_event_key"] for item in rows}
    manifest = {
        f"{item['source_event_key']}|{item['expected_row_hash']}"
        for item in rows
    }
    _validate_hash(
        _hash_lines(event_keys),
        EXPECTED_DECISION_EVENT_KEYS_HASH,
        "decision event-key manifest",
    )
    _validate_hash(
        _hash_lines(manifest),
        EXPECTED_DECISION_OBSERVATION_MANIFEST_HASH,
        "decision observation manifest",
    )
    _validate_hash(
        _canonical_hash(rows),
        EXPECTED_DECISION_PAYLOAD_HASH,
        "decision payload",
    )
    return rows


def load_review_rows(
    connection: sqlite3.Connection,
) -> list[dict[str, Any]]:
    rows = []
    for instrument_id, event_key, row_hash in REVIEW_SPECS:
        observation = _single_current_observation(
            connection, instrument_id, event_key
        )
        if observation["row_hash"] != row_hash:
            raise RuntimeError(
                "manual-review observation row hash drifted: " + event_key
            )
        if (
            observation.get("resolution_state")
            != "official_archive_unavailable"
        ):
            raise RuntimeError(
                "manual-review event is no longer an archive gap: "
                + event_key
            )
        rows.append(observation)
    if len(rows) != 10 or len({
        item["source_event_key"] for item in rows
    }) != 10:
        raise RuntimeError("manual-review manifest must contain ten events")
    event_keys = {item["source_event_key"] for item in rows}
    manifest = {
        f"{item['source_event_key']}|{item['row_hash']}"
        for item in rows
    }
    _validate_hash(
        _hash_lines(event_keys),
        EXPECTED_REVIEW_EVENT_KEYS_HASH,
        "manual-review event-key manifest",
    )
    _validate_hash(
        _hash_lines(manifest),
        EXPECTED_REVIEW_OBSERVATION_MANIFEST_HASH,
        "manual-review observation manifest",
    )
    return sorted(
        rows,
        key=lambda item: (
            item["instrument_id"],
            item["source_event_key"],
        ),
    )


def _hash_query(
    connection: sqlite3.Connection,
    query: str,
    params: tuple[Any, ...] = (),
) -> dict[str, Any]:
    rows = [dict(row) for row in connection.execute(query, params)]
    return {"rows": len(rows), "sha256": _canonical_hash(rows)}


def immutable_snapshot(
    connection: sqlite3.Connection,
    decisions: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    event_keys = sorted(item["source_event_key"] for item in decisions)
    instrument_ids = sorted({
        item["instrument_id"] for item in decisions
    })
    event_marks = ",".join("?" for _ in event_keys)
    instrument_marks = ",".join("?" for _ in instrument_ids)
    snapshot = {
        "cninfo_observations": _hash_query(
            connection,
            f"""
            SELECT *
            FROM corporate_action_observations
            WHERE source='cninfo'
              AND source_event_key IN ({event_marks})
            ORDER BY id
            """,
            tuple(event_keys),
        ),
        "tdx_rows": _hash_query(
            connection,
            f"""
            SELECT *
            FROM adjustment_factors_tdx
            WHERE instrument_id IN ({instrument_marks})
            ORDER BY id
            """,
            tuple(instrument_ids),
        ),
    }
    for table in PRODUCTION_FACTOR_TABLES:
        snapshot[table] = _hash_query(
            connection,
            f"""
            SELECT *
            FROM {table}
            WHERE instrument_id IN ({instrument_marks})
            ORDER BY id
            """,
            tuple(instrument_ids),
        )
    snapshot["adjustment_factor_series_status"] = _hash_query(
        connection,
        """
        SELECT *
        FROM adjustment_factor_series_status
        ORDER BY series_version
        """,
    )
    return snapshot


def _validate_write_database_path(database_path: Path) -> None:
    from utils import config_manager

    requested = database_path.expanduser().resolve()
    if requested != DEFAULT_DATABASE.resolve():
        raise ValueError("--write is restricted to the configured quotes.db")
    configured = Path(config_manager.get_nested(
        "database_config.db_path", ""
    )).expanduser().resolve()
    if configured != requested:
        raise RuntimeError(
            f"configured database path mismatch: {configured} != {requested}"
        )


async def apply_decisions(
    decisions: list[dict[str, Any]],
    database_path: Path,
) -> list[dict[str, Any]]:
    os.chdir(ROOT_DIR)
    _validate_write_database_path(database_path)
    from data_manager import DataManager

    manager = DataManager()
    results = []
    for sequence, decision in enumerate(decisions, start=1):
        print(
            f"[{sequence}/{len(decisions)}] applying "
            f"{decision['instrument_id']} "
            f"{decision['terminal_reason']} "
            f"{decision['source_event_key']}",
            flush=True,
        )
        result = (
            await manager.review_cninfo_corporate_action_terminal_disposition(
                dict(decision)
            )
        )
        results.append({
            "sequence": sequence,
            "instrument_id": decision["instrument_id"],
            "source_event_key": decision["source_event_key"],
            "terminal_reason": decision["terminal_reason"],
            "review_id": result["review"]["review_id"],
            "resolution_state": result["resolution_state"][
                "resolution_state"
            ],
        })
        print(
            f"[{sequence}/{len(decisions)}] applied "
            f"review_id={result['review']['review_id']}",
            flush=True,
        )
    return results


def audit_written_decisions(
    connection: sqlite3.Connection,
    decisions: list[dict[str, Any]],
) -> dict[str, Any]:
    event_keys = sorted(item["source_event_key"] for item in decisions)
    placeholders = ",".join("?" for _ in event_keys)
    state_rows = connection.execute(
        f"""
        SELECT instrument_id, source_event_key, resolution_state,
               is_terminal, factor_blocking
        FROM corporate_action_resolution_states
        WHERE source_event_key IN ({placeholders})
        ORDER BY source_event_key
        """,
        event_keys,
    ).fetchall()
    if len(state_rows) != 62:
        raise RuntimeError("state audit did not find all 62 decisions")
    expected_by_key = {
        item["source_event_key"]: item["terminal_reason"]
        for item in decisions
    }
    for row in state_rows:
        if (
            row["resolution_state"]
            != expected_by_key[row["source_event_key"]]
            or not bool(row["is_terminal"])
            or bool(row["factor_blocking"])
        ):
            raise RuntimeError(
                "terminal state audit failed: " + row["source_event_key"]
            )

    review_rows = connection.execute(
        f"""
        SELECT r.*
        FROM corporate_action_resolution_reviews AS r
        WHERE r.source_event_key IN ({placeholders})
          AND r.id=(
              SELECT latest.id
              FROM corporate_action_resolution_reviews AS latest
              WHERE latest.source_event_key=r.source_event_key
              ORDER BY latest.updated_at DESC, latest.id DESC
              LIMIT 1
          )
        ORDER BY r.source_event_key
        """,
        event_keys,
    ).fetchall()
    if len(review_rows) != 62:
        raise RuntimeError("latest review audit did not find all 62 decisions")
    reason_counts = Counter()
    for row in review_rows:
        payload = json.loads(row["review_payload_json"])
        terminal_reason = str(payload.get("terminal_reason") or "")
        if (
            row["reviewer"] != REVIEWER
            or row["decision"] != "rejected"
            or row["effective_date"] is not None
            or row["analysis_id"] is not None
            or payload.get("effective_date_intentionally_absent") is not True
            or terminal_reason
            != expected_by_key[row["source_event_key"]]
        ):
            raise RuntimeError(
                "latest review lineage audit failed: "
                + row["source_event_key"]
            )
        reason_counts[terminal_reason] += 1
    if dict(reason_counts) != {
        "non_effective": 17,
        "pre_listing": 45,
    }:
        raise RuntimeError(
            f"unexpected terminal reason counts: {dict(reason_counts)}"
        )
    return {
        "review_count": len(review_rows),
        "state_count": len(state_rows),
        "terminal_reason_counts": dict(reason_counts),
        "remaining_blocker_count": 0,
    }


def _reference_dates(row: Mapping[str, Any]) -> list[date]:
    operational = [
        parsed
        for field in ("record_date", "pay_date", "share_arrival_date")
        if (parsed := _date(row.get(field))) is not None
    ]
    if operational:
        return sorted(set(operational))
    announcement_date = _date(row.get("announcement_date"))
    return [announcement_date] if announcement_date else []


def _nearest_tdx(
    rows: list[dict[str, Any]],
    anchors: list[date],
    *,
    limit: int = 3,
) -> list[dict[str, Any]]:
    if not anchors:
        return []
    ranked = []
    for row in rows:
        ex_date = _date(row.get("ex_date"))
        if ex_date is None:
            continue
        gap = min(abs((ex_date - anchor).days) for anchor in anchors)
        ranked.append({**row, "nearest_gap_days": gap})
    return sorted(
        ranked,
        key=lambda item: (
            int(item["nearest_gap_days"]),
            str(item.get("ex_date") or ""),
            int(item.get("id") or 0),
        ),
    )[:limit]


def build_workbook_frames(
    rows: list[dict[str, Any]],
    tdx_by_instrument: dict[str, list[dict[str, Any]]],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    main_rows = []
    tdx_rows = []
    for row in rows:
        instrument_id = str(row["instrument_id"])
        candidates = tdx_by_instrument.get(instrument_id, [])
        match = match_cninfo_archive_tdx_date(
            row, candidates, field_tolerance=0.0001
        )
        if match.get("matched"):
            raise RuntimeError(
                "manual-review event unexpectedly gained a unique TDX match: "
                + row["source_event_key"]
            )
        anchors = _reference_dates(row)
        nearest = _nearest_tdx(candidates, anchors)
        nearest_one = nearest[0] if nearest else {}
        main_rows.append({
            "序号": len(main_rows) + 1,
            "证券代码": instrument_id,
            "证券名称": row.get("instrument_name") or "",
            "事项键": row["source_event_key"],
            "财年/期间": row.get("fiscal_period") or "",
            "上市日期": _date_text(row.get("listed_date")),
            "CNInfo公告日": _date_text(row.get("announcement_date")),
            "CNInfo登记日": _date_text(row.get("record_date")),
            "CNInfo派息日": _date_text(row.get("pay_date")),
            "CNInfo股份到账日": _date_text(
                row.get("share_arrival_date")
            ),
            "CNInfo现金/10股": _per_ten(
                row.get("cash_dividend_per_share")
            ),
            "CNInfo送股/10股": _per_ten(
                row.get("bonus_shares_per_share")
            ),
            "CNInfo转增/10股": _per_ten(
                row.get("capitalization_shares_per_share")
            ),
            "CNInfo配股/10股": _per_ten(
                row.get("rights_shares_per_share")
            ),
            "CNInfo配股价": row.get("rights_price"),
            "CNInfo事项说明": row.get("description") or "",
            "当前治理状态": row.get("resolution_state") or "",
            "不能自动核准原因": (
                "历史实施公告不可得；现有日期均在上市后，且未找到"
                "唯一、经济条款一致的TDX日期旁证。系统不能猜测除权日。"
            ),
            "TDX自动匹配结果": match.get("reason") or "",
            "最近TDX记录ID": nearest_one.get("id") or "",
            "最近TDX除权日": _date_text(nearest_one.get("ex_date")),
            "最近TDX距参考日/天": nearest_one.get(
                "nearest_gap_days", ""
            ),
            "最近TDX现金/10股": nearest_one.get("fenhong"),
            "最近TDX送转/10股": nearest_one.get("songzhuangu"),
            "最近TDX配股/10股": nearest_one.get("peigu"),
            "最近TDX配股价": nearest_one.get("peigujia"),
            "建议人工判断": (
                "确认真实除权日；若事项只针对上市前/老股东或不影响"
                "上市流通价格，可标记不参与复权。"
            ),
            "用户决定": "",
            "核准除权日": "",
            "factor_effect": "",
            "用户说明": "",
        })
        for rank, candidate in enumerate(nearest, start=1):
            tdx_rows.append({
                "事项键": row["source_event_key"],
                "证券代码": instrument_id,
                "证券名称": row.get("instrument_name") or "",
                "CNInfo参考日期": "、".join(
                    item.isoformat() for item in anchors
                ),
                "距离排序": rank,
                "TDX记录ID": candidate.get("id"),
                "TDX除权日": _date_text(candidate.get("ex_date")),
                "距最近参考日/天": candidate.get("nearest_gap_days"),
                "TDX现金/10股": candidate.get("fenhong"),
                "TDX送转/10股": candidate.get("songzhuangu"),
                "TDX配股/10股": candidate.get("peigu"),
                "TDX配股价": candidate.get("peigujia"),
                "TDX因子": candidate.get("factor"),
                "TDX校验状态": candidate.get("validation_result"),
            })

    summary = pd.DataFrame([
        {"项目": "待人工审核事项", "数值": 10, "说明": "均为上市后且缺可靠除权日"},
        {"项目": "已确认上市前事项", "数值": 45, "说明": "42条公告锚点 + 3条明确实施日期"},
        {"项目": "已确认无经济影响事项", "数值": 17, "说明": "利润滚存、归老股东/发起人等说明性记录"},
        {"项目": "重新下载公告", "数值": 0, "说明": "network_access=false"},
        {"项目": "OCR运行", "数值": 0, "说明": "run_ocr=false"},
        {"项目": "LLM调用", "数值": 0, "说明": "llm_invocations=0"},
    ])
    fields = pd.DataFrame([
        {
            "字段": "用户决定",
            "填写说明": "建议填写：通过 / 修改后通过 / 不参与复权 / 继续补证据",
        },
        {
            "字段": "核准除权日",
            "填写说明": "仅在确认真实上市市场生效交易日后填写，格式YYYY-MM-DD",
        },
        {
            "字段": "factor_effect",
            "填写说明": "normal / none / official_reference_price",
        },
        {
            "字段": "用户说明",
            "填写说明": "记录判断依据、特殊受益范围、停牌或上市前情况",
        },
        {
            "字段": "TDX邻近记录",
            "填写说明": "仅作日期和事项身份旁证，不覆盖CNInfo经济数字",
        },
    ])
    return (
        summary,
        pd.DataFrame(main_rows),
        pd.DataFrame(tdx_rows),
        fields,
    )


def write_workbook(
    path: Path,
    frames: tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame],
) -> dict[str, Any]:
    path.parent.mkdir(parents=True, exist_ok=True)
    sheet_names = (
        "00_总览",
        "01_10条待审核",
        "02_TDX邻近事项",
        "03_字段说明",
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, frame in zip(sheet_names, frames):
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
        for column_index, column_cells in enumerate(
            sheet.columns, start=1
        ):
            values = [
                str(cell.value or "") for cell in column_cells[:40]
            ]
            width = min(48, max(10, max(map(len, values), default=10) + 2))
            sheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width
    review_sheet = workbook["01_10条待审核"]
    headers = {
        str(cell.value): cell.column
        for cell in review_sheet[1]
    }
    for name in ("用户决定", "核准除权日", "factor_effect", "用户说明"):
        column = headers[name]
        for row_index in range(2, review_sheet.max_row + 1):
            review_sheet.cell(row_index, column).fill = input_fill
    workbook.save(path)
    workbook.close()
    return {
        "path": str(path.resolve()),
        "sha256": sha256(path.read_bytes()).hexdigest(),
        "bytes": path.stat().st_size,
        "sheet_rows": {
            name: len(frame)
            for name, frame in zip(sheet_names, frames)
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--write",
        action="store_true",
        help="Persist the fixed 62 terminal decisions.",
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=DEFAULT_DATABASE,
    )
    parser.add_argument(
        "--report",
        type=Path,
        default=DEFAULT_REPORT,
    )
    args = parser.parse_args()

    connection = _connect_read_only(args.database)
    try:
        decisions = build_decisions(connection)
        review_rows = load_review_rows(connection)
        before = immutable_snapshot(connection, decisions)
        tdx = _load_tdx(
            connection,
            [row["instrument_id"] for row in review_rows],
        )
    finally:
        connection.close()

    result: dict[str, Any] = {
        "status": "validated_preview",
        "write_requested": bool(args.write),
        "decision_count": len(decisions),
        "decision_counts": dict(Counter(
            item["terminal_reason"] for item in decisions
        )),
        "decision_event_keys_hash": _hash_lines(
            item["source_event_key"] for item in decisions
        ),
        "decision_observation_manifest_hash": _hash_lines(
            f"{item['source_event_key']}|{item['expected_row_hash']}"
            for item in decisions
        ),
        "decision_payload_hash": _canonical_hash(decisions),
        "review_event_keys_hash": _hash_lines(
            item["source_event_key"] for item in review_rows
        ),
        "review_observation_manifest_hash": _hash_lines(
            f"{item['source_event_key']}|{item['row_hash']}"
            for item in review_rows
        ),
        "manual_review_count": len(review_rows),
        "network_access": False,
        "document_downloads": 0,
        "ocr_invocations": 0,
        "llm_invocations": 0,
        "immutable_snapshot_before": before,
    }
    if args.write:
        result["writes"] = asyncio.run(
            apply_decisions(decisions, args.database)
        )
        connection = _connect_read_only(args.database)
        try:
            result["write_audit"] = audit_written_decisions(
                connection, decisions
            )
            after = immutable_snapshot(connection, decisions)
            if after != before:
                raise RuntimeError(
                    "raw CNInfo, TDX, or production factor data changed"
                )
            result["immutable_snapshot_after"] = after
            review_rows = load_review_rows(connection)
            tdx = _load_tdx(
                connection,
                [row["instrument_id"] for row in review_rows],
            )
        finally:
            connection.close()
        result["report"] = write_workbook(
            args.report,
            build_workbook_frames(review_rows, tdx),
        )
        result["status"] = "success"
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
