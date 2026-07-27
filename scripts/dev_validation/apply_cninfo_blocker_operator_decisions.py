#!/usr/bin/env python3
"""Preview or apply the fixed 55 CNInfo blocker-workbook decisions."""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import sqlite3
import sys
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


ROOT_DIR = Path(__file__).resolve().parents[2]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))


DEFAULT_DATABASE = ROOT_DIR / "data/quotes.db"
DEFAULT_WORKBOOK = (
    ROOT_DIR
    / "data/reports/cninfo_corporate_action_blockers_63_review_20260727.xlsx"
)
REVIEWER = "operator_cninfo_blocker_workbook_20260727"
EXPECTED_ALL_EVENT_KEYS_HASH = (
    "b653c9ca3a678a95768e147d3fbfd428ec87a68f7c03d78aacabc0e1db82d596"
)
EXPECTED_APPROVED_EVENT_KEYS_HASH = (
    "5780b23fbb1755a66c8b1d3b2cd02b9aa28673bd478d9c180d78a4c455e49dac"
)
EXPECTED_OBSERVATION_MANIFEST_HASH = (
    "725bdcafd91954fa8b1172b0b96a23700d8824517c457e12df4f2ccfa5830500"
)
EXPECTED_DECISION_PAYLOAD_HASH = (
    "f02904c4653713b203c948f38d0f82da8ec133c6ebab796c2ad206ccfb6d938c"
)

ECONOMIC_FIELDS = (
    "cash_dividend_per_share",
    "bonus_shares_per_share",
    "capitalization_shares_per_share",
    "rights_shares_per_share",
    "rights_price",
)
PRODUCTION_FACTOR_TABLES = (
    "adjustment_factor_observations",
    "adjustment_factors",
    "adjustment_factors_canonical",
    "adjustment_factor_instrument_status",
)

# instrument, event key, row hash, analysis, announcement, effective date,
# date basis, factor effect
DECISION_SPECS: tuple[tuple[Any, ...], ...] = (
    ("000155.SZ", "c9f3d09836843bd3112c9b2d4a53c7f22abcb26e504646dd0eed41e5b1d87875", "bfdd6fb4a5f61a15a3a44fe57b4dcac0e25c233e0e568326ca3d3dd34cd13f04", 547, "1202868053", "2016-12-13", "用户核准的官方record_date", "none"),
    ("000410.SZ", "abaf5732d6b62168d0a4e8d9ae11895762a682437ad8620804d18a7880abce01", "6bb27e2fefe1c556b83057879a97f614eca348dbd0821332a60712996f2fb9af", 665, "1207169746", "2019-12-23", "用户核准的官方listing_date", "none"),
    ("000430.SZ", "4b0c92a1f5738e22f0bde525663c7bf893b5b90a22acc38e41de54bd052b4d34", "c57c5ac06b068ed599893e969737b65b351f55b9a5aa78c2af346e92934e2509", 553, "1224901404", "2025-12-29", "官方调整后首个交易日", "official_reference_price"),
    ("000520.SZ", "061473592b6b5b07c7be4ce2269ff9e518b171a8f7c50291a78c6489010476c2", "8107ea39a677d349edeb55e7e84469e384ddd3d6b661bef9b3239cc1df0ea27e", 555, "1200074068", "2014-07-30", "用户核准的官方share_arrival_date", "none"),
    ("000523.SZ", "00531a5d7532c64250ca1011a91acdc0d98f6d75c8cba1cf86268005e131fdb0", "5ceb92163d6326708e91b86f5aa520f729289fb9443fe895c4c92d325c21fdec", 559, "1211779449", "2021-12-09", "用户核准的官方listing_date", "none"),
    ("000557.SZ", "b061f9e888b081dbef432a199393dfcd4f064544210ed43c07060186be2817d0", "1eaa73bb5f1db2e12f35d2e6e38c4f34b57ca095a2e9673283c0452f2dabd702", 562, "20935416", "2007-02-26", "用户指定TDX XDXR除权交易日", "normal"),
    ("000557.SZ", "b3218cf58c2eac9cfcaa46418642e1a5b8bff42f24be771788f961c5779b7786", "b1ada9ae8871adb8b5c2c2d017d79ce23587a79bef5e33371f476aeb5a18556f", 583, "47289248", "2008-12-17", "用户核准的官方share_arrival_date", "normal"),
    ("000595.SZ", "6db3f64f3e3b99f51ac3f4345a88f7228236ad865edfb38d12e605b1078a7526", "30557788831c697bacfa439ffc74cfea00b0e99c2aa34eac54eb146ce44efadb", 631, "1208793565", "2020-11-30", "用户核准的官方listing_date", "none"),
    ("000656.SZ", "b8a5535ff306d3c83cae66c0546eef55805d831ffe4c64c38b8b746c5085e0b5", "d935b24186bc64d0636caa901d4c325971585381395e26dcc5e8eec563d8ddbe", 574, "1224645547", "2025-09-12", "用户核准的官方resumption_date", "normal"),
    ("000691.SZ", "0e4cf241f065a3fb547f5a4e1d6b1d47f9c03c312977023ed2eb9fb64df55bfe", "15f521d203bad7a87be9c5636d0b4aed776cf438d8298bf5c423f76834c18a16", 653, "1224902023", "2026-01-05", "用户核准的官方resumption_date", "normal"),
    ("000697.SZ", "4c6b0f737e290ffea12f183203d170ecb46d1866a835f3a5330c806adf995d01", "c54d2d298a926624ee7b27af2865d87456b11c2784ed6a8a19989139c8bfa0b9", 650, "1224831705", "2025-11-28", "用户核准的官方ex_date", "normal"),
    ("000711.SZ", "37534bea29f4b946a9e285c307082c3cacbdcd89461779240e56a0019ac523bf", "b9137a9dab1c786fb675d41b72b1ced41781bf16efe3593142341884e8197614", 610, "1218581599", "2023-12-18", "用户核准的官方resumption_date", "normal"),
    ("000751.SZ", "767658ee481aa0e1cfd089d99ba8bf2cbf202292138cafd94e42cb79d3ef4860", "36a649bdba57efe81f078a906c2a14cb493473f80e0469b439fe8bc3cb0ef622", 677, "63415685", "2013-12-31", "用户核准的官方share_arrival_date", "normal"),
    ("000792.SZ", "f815d9a7870e481cc217de5bc1b685bdf53e8868fc1db6c6e928a3c845d20a31", "a500f385471a5add73e64e92cc80889109645fae58c904cf64e01ae8b81731bd", 586, "1207399015", "2020-03-31", "用户核准的官方listing_date", "none"),
    ("000793.SZ", "52d60258dd5282141e059f95eeb10fb919b168b341b1bc5ee520092619d77b5b", "b56868fff51c2867d820c18ec2b7bab7ca946c357326079ab0d24a41a47d39ee", 703, "1225380075", "2026-06-22", "官方调整后首个交易日", "official_reference_price"),
    ("000908.SZ", "472b070d3e97113e7f8694d6a4e0b290748e935d4ba1c8fdb5ad23219e640e0d", "9b788c7a2989e3f387c653faa5bf9371f800efaa7687f929ce8ac53248c300b4", 616, "1224993850", "2026-03-11", "用户核准的官方resumption_date", "normal"),
    ("000912.SZ", "5490b5e2329c63d8f7bbeaef0018142afbc6891f3c13a7b6b937ad2270107d06", "a659fa678f09961e539d8e54c26880dc90c6d872b88d4efa38d7094000e8e939", 695, "1205239664", "2018-07-27", "用户核准的官方share_arrival_date", "normal"),
    ("000972.SZ", "ea770f6160847d86110e880b80519442eb96398c2150f3fec3445228e118ec0e", "033cd693e7968507a5f568684ea1347939ee1fcf86053cf700a73149c3f2d4d4", 615, "62133369", "2013-02-22", "用户核准的官方record_date", "normal"),
    ("000980.SZ", "92103259a13e834e4e453e6842c57ca356c84c36d4cd4e7272a40cdafcbd6f31", "d9e9ffd7223f63bfaa6657efe0b0e2181de753b532d68447538c80c74f618440", 717, "1211847799", "2021-12-15", "用户核准的官方listing_date", "none"),
    ("002072.SZ", "6bb47db200900939b251baf1c83b9fde6b471eba1301fc34faac676ddd4c0af5", "ba208c4a1a8c1987807f524854875628677d120903eb93b9d766604e701ee1aa", 725, "1211946446", "2021-12-24", "用户核准的官方listing_date", "none"),
    ("002086.SZ", "c55001f4817b36cc0a64572eb8270f765487d7b3854955b5606017bb3440631f", "287f722cd4cc36274bd7418e9125296a20c4966ab1cdaf32e5706d6ff5196ca6", 637, "1218707072", "2023-12-29", "用户核准的官方resumption_date", "normal"),
    ("002122.SZ", "c4f37dfd314de37bbd1725aedea4bce52a6b1b7d56d6b37fe58038a2e9b52609", "943b1ee2f154e20b7b3a134e992ad64c9db410b77e7359eff58c0f0a1c78fa54", 633, "1215367221", "2022-12-22", "用户核准的官方resumption_date", "normal"),
    ("002131.SZ", "fb50912e2341436486653d7fc5880fa7b08aef83fe8d2af2c0fb492a388094a9", "72b1f4023c93597077c07e7454d96e86fdb19b8eba5fa91689edbbb866686b4e", 739, "1204199768", "2017-12-15", "用户核准的官方share_arrival_date", "normal"),
    ("002157.SZ", "1dde12cd91fb048fbbb7057320088d7f48109d3fc965b4aaac4ee99f9db5f735", "475efe7a167e17cf48fbcbcd649f6ae56bc9970cad005207a206b6c42aa6b5c9", 632, "1218498400", "2023-12-08", "用户核准的官方resumption_date", "normal"),
    ("002175.SZ", "05b38afda50a5e48a9c07bfb02d832f729cda387bd68367ed3c8ebe41526e96e", "d50acd4002fd214becf393bcd36e68cc8e0a92127c30096bb16d1e28503f0ccc", 742, "1211970924", "2021-12-27", "用户核准的官方listing_date", "none"),
    ("002192.SZ", "2a9c8213085fae289667264b5c467ed8bf1cc693125c33ef6667864055898a8a", "6181e72027f271bf71c955a0e0e88d0c7ee3e4d62cd9d9b66c8fc5942c08c1c7", 630, "1205194969", "2018-10-17", "用户核准的官方pay_date", "normal"),
    ("002200.SZ", "b6136cbbdf639b0ca34a0c95b88445b6fc26761a5083ecb6828d9c56edacc8d0", "2d1eef8ce4f28cbae1fe922bf9efe828fa298ce05f522131affd87f265ac0582", 749, "1224869457", "2025-12-12", "官方调整后首个交易日", "official_reference_price"),
    ("002210.SZ", "feafc0d6e9aba233b41de2e34ac40270b10b78b0390e51d461eac4acb45e8f22", "7bae34450d1ca775248039f9bb13c28a8ef15cb21fc8325ed2f311da4a829611", 634, "1208971425", "2020-12-30", "用户核准的官方listing_date", "none"),
    ("002217.SZ", "40a7ffc2e61508f55062014e6eae13b950d3c817ea17d6d846bd42032c9146af", "d14ec5f746683bf5706f65811315a96ee8657426d9fb3ec306cb2e1ec56c9d4b", 646, "1222166633", "2024-12-31", "用户核准的官方resumption_date", "normal"),
    ("002310.SZ", "d2aedecd155e903d53674b1a3bad1d5afd63a3be12f33d50d52c9405edb600e5", "acc98504181780383da1f910ae8169ad5d440a4b45b5464f28d9d0b90cb213f2", 741, "1222115435", "2024-12-30", "用户核准的官方resumption_date", "normal"),
    ("002354.SZ", "f927d515201e0e52cfaf444826741f10c529c3a4b0d7386819e09f6f2b150ebd", "27212ef36c48f12d6858f20676b0e217c8b732c717d02177d54d4a5a1f27837f", 728, "1208829627", "2020-12-08", "用户核准的官方share_arrival_date", "normal"),
    ("002366.SZ", "377ef018caf523a9feea94c4b6b6180d20aedf1270d89167a18af8d40d57dd0b", "8c4a6e795af3aca6d77283787c760abc70688b71f25e6ddbdf49ec1e71a26f2a", 639, "1215418604", "2022-12-22", "用户核准的官方ex_date", "normal"),
    ("002427.SZ", "775898ee3854c5a69fe3d71fea839978501e6b2e84251e8e48206d76b83f83fa", "c35c737e4ec861e103edc46c542dbfeefeea4a52559e75f97d19d67f1132e425", 640, "1215282461", "2022-12-12", "用户核准的官方resumption_date", "normal"),
    ("002445.SZ", "72c1ecb0150274ec86ac69e787b2012e364aa28903e22994a804d4cfa677007c", "6d12adbd57713b0d83c590f1fe41ef0123ac714885d5608cf14089bad668ad15", 638, "1208994482", "2020-12-31", "用户核准的官方listing_date", "none"),
    ("002482.SZ", "7ba93491550396719c221e98e66e6bead263d20a817eba481465d809e25345d1", "4b293f67e25710ecbc281cf5191810a8aad777163ad18d419e0ea11bb6ad8b8b", 730, "1218609774", "2023-12-20", "用户核准的官方resumption_date", "normal"),
    ("002501.SZ", "6bcc7e3e15aefb4c7127c80387a2e8748fcd9b0492af8689544a1e58becabf0c", "7cbbe07c924aa9e4b6a3fb147a7f96a11023fead3d5d8ef677ab122e3d9eb5be", 641, "1209008336", "2020-12-31", "用户核准的官方listing_date", "none"),
    ("002608.SZ", "61b4c801628a62d2cf0da09f749512732a10d26722eec8d098befa98c129070a", "60a834da51406ef74ff7c51003658d967b8d4dd47ea971d9016147305cff7ce0", 644, "1202965913", "2016-12-28", "用户核准的官方share_arrival_date", "none"),
    ("002713.SZ", "729b103b9896c9111f598a2426f67e8bb98fff8cc957f0939948c478b97714d4", "6b11c252363759c22b0217e4cccc063cba2565d2fc73c0f55d55da5050fb08dc", 656, "1224906744", "2025-12-30", "用户核准的官方ex_date", "normal"),
    ("002716.SZ", "90c3cd416df81a5373cc029d5b67f4a5a351f39924c2cd4327305abfff7d365e", "d085e7864419f1d7a3317b65901c629eb80df76f1bee25eacc82f7b4ae919ba6", 729, "1208971926", "2020-12-29", "用户核准的官方listing_date", "none"),
    ("002822.SZ", "c1b35b94c9d212b8637405a48b7e42c39fe6b14fa7afff517d55c00ec39028e0", "af33c4c84201ad1af69d839f560e62fdd0ec557e62513ef07126764b933d1018", 764, "1224906788", "2025-12-30", "用户核准的官方ex_date", "normal"),
    ("300008.SZ", "7cfa6aefb91d6a5d7f09f71baf7d6a16ded103de16b1ee37fc6ec0f328ee5c00", "8f63f625f332d96bb01d6c2e1e4fb259a43e185543f37020367283b404bf89ad", 647, "1208928813", "2020-12-23", "用户核准的官方share_arrival_date", "none"),
    ("300071.SZ", "c699faa8ec555d37a76ce2b814d2c028e5be248af5cdf891ca0698ad9f0e7687", "a334289cf36f1363fab7f8dc2508a0580382db7aa5cea9de6608d94373aa590d", 754, "1212031481", "2021-12-30", "用户核准的官方listing_date", "normal"),
    ("300125.SZ", "121b35c36e5703b91c5b88a7f7b8708f40e29831e90fd7d9e839a6efa8b0caeb", "0113a861ef58fffa901d97192fb028dbe13329bc4dca8242c6d743b3bbc6fab1", 661, "1224901802", "2025-12-29", "官方调整后首个交易日", "official_reference_price"),
    ("300180.SZ", "b4fb6a03ae358f0e6b17f69b1a4541a125a85bb2d951586eee0a39da852c07b9", "e0b38988b9a9c70bead35b1b4e7994a5999478d2f207f70349f65666797b127d", 765, "1207932072", "2020-06-16", "用户核准的官方listing_date", "normal"),
    ("300256.SZ", "2db66101f4ceaf06c270c2353af5bf2af2a4a4dfcf70f59ed85e1c5fd23d2ba8", "a8c3d07fa567e07f5cc1bf28c176529d4e778ebcebe7b8a707467fd24f0b1659", 664, "1214388605", "2022-08-31", "用户核准的官方resumption_date", "normal"),
    ("300278.SZ", "1dae97062a0e5fb90db385e268f31082ab51d9e13b4aeaa9342fafc88913f6f6", "891f799d9d6153d1312792ebdc3345afb35834325c50353c38014cde04e044f3", 662, "1212033350", "2021-12-31", "用户核准的官方resumption_date", "normal"),
    ("300370.SZ", "30794f4afb765e12e423a868121385ab66e2a0caab92b6cc72a6bf5651056526", "366084c464eee77ce59a7f2182c3859801e42d42c3a7902dcdea18d82ff3ed67", 678, "1215399349", "2022-12-20", "官方调整后首个交易日", "official_reference_price"),
    ("300506.SZ", "1505cd09a94602716b50f42fbe70cafd722ba760ac3cdfff915d9a2d2d2cf405", "5e8d6efa2447083f9bfd68e4877e984aaf4f0a995696d5d57508526c069c9e8e", 757, "1224877791", "2025-12-22", "用户核准的官方resumption_date", "normal"),
    ("600280.SH", "8df255655339ec6a9e68e77969561827f328f8f8e8052c35a556162625da324e", "a1aa414c6928494996ca91a8dd86b1e8a2503654a6b5e4c79993b251b38af315", 762, "506911", "2001-06-26", "用户指定TDX XDXR除权交易日", "none"),
    ("600295.SH", "1f3a844683e08632b6171d2465d60a092fe469d036aaf19511f208cb6a3728f8", "969082aadffa8ca3cc06f52c3a80730d4fcfe4d967ed52d6a07256d8b06200c3", 684, "507659", "2001-07-06", "用户核准的官方ex_date", "none"),
    ("600315.SH", "09a8f88dfee8e99902b90a68f04f5a61df8cc1eeb4ebd87ff56f94940a8848b4", "197849c2a97e40330fc715f0b7314d519c2f52ac40fc3771295c4997ae927f50", 685, "508410", "2000-07-31", "用户核准的官方record_date", "none"),
    ("600399.SH", "63a247667cf8f77314249c4dda580658fbb050604720417dc3d6f2673fd91ab7", "8f294ba68704688458fb52254413ee86ebf72ee46b39b3e6098e5920c7dc1281", 697, "1205682049", "2018-12-28", "用户核准的官方resumption_date", "normal"),
    ("600572.SH", "fa5c0b22c5fdfbfee8e431ef20be1df73ad8b16546f4d507a7db889b980d3edf", "5e1020c9ab63ed81e461f45cb50c4ecd8f2efe625422ff483df3a4285ff17340", 707, "14199453", "2004-03-29", "用户核准的官方record_date", "none"),
    ("600733.SH", "dcfe8ecbf8d55058f450c838755c67d1b80b69b67a1e0080924707690f590db0", "a6ba3bede60a1afae9014fd7022959bab64c6f7fd28eab5234c942e78e26a632", 753, "1205432898", "2018-09-19", "用户指定TDX XDXR除权交易日", "normal"),
    ("600828.SH", "8c40611c0c717bb581191e21d640e9e552107e76340d04a6d7845e079b99ffe6", "ee937569476d8aae40e09f2472da48dbedeb9635c5c05bcbcf804859bc1e3875", 795, "1206276658", "2019-05-24", "用户核准的官方listing_date", "normal"),
)

DEFERRED_REASONS = {
    "2f0184e8c91fb9527402445521b488e13a47a26a6b8bbb00ea98677612cda67d": (
        "同一公告存在冲突登记日且无最终除权或上市日期"
    ),
    "cba34363197b4120854fa95a291f7c1a1b80697bc7052c2b3a39f62d5fcb204a": "无可用官方证据",
    "78aa955b66f4c09a1d46b8ff4b9e067bf031b2cf0e1f00dea4c393f6f345bfaf": "无可用官方证据",
    "283fa0fa2c61e7fc1b6275d89c782a3e840fba94adf4a1a93416ec5788175065": "无可用官方证据",
    "fc72dec38e8cf3177d43716ef2efe0a796208d4b6b82917270a5f6039c6444bb": "无可用官方证据",
    "d03958e8cfda4911156de3d7645cda8eb7ec65b56d795258069b2ea279195fe3": "无可用官方证据",
    "1b12e1cd4a497920651b3a0fb000e1200af24acb3cdd18218b9ca1eecdc0b0c4": "无可用官方证据",
    "44eb085369be9a4b74d11e28d9e7c0a77921911871827679c51125ebd1f7c141": "仅有方案阶段公告",
}
TDX_DATE_ROWS = {
    "8df255655339ec6a9e68e77969561827f328f8f8e8052c35a556162625da324e": 6560,
    "b061f9e888b081dbef432a199393dfcd4f064544210ed43c07060186be2817d0": 30597,
    "dcfe8ecbf8d55058f450c838755c67d1b80b69b67a1e0080924707690f590db0": 13807,
}
OFFICIAL_FACTOR_REFERENCES = {
    "000430.SZ": (7.90, 6.87),
    "000793.SZ": (2.63, 2.51),
    "002200.SZ": (7.82, 6.94),
    "300125.SZ": (7.86, 6.22),
    "300370.SZ": (2.66, 2.64),
}


def _hash_lines(values: Iterable[str]) -> str:
    return sha256("\n".join(sorted(values)).encode("utf-8")).hexdigest()


def _canonical_hash(value: Any) -> str:
    return sha256(json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")).hexdigest()


def _workbook_event_keys(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["01_63条逐项"]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        key_index = headers.index("事项键")
        keys = {
            str(row[key_index]).strip()
            for row in rows
            if row[key_index] is not None and str(row[key_index]).strip()
        }
    finally:
        workbook.close()
    return keys


def _spec_rows() -> list[dict[str, Any]]:
    field_names = (
        "instrument_id",
        "source_event_key",
        "expected_row_hash",
        "analysis_id",
        "announcement_id",
        "effective_date",
        "date_basis",
        "factor_effect",
    )
    return [dict(zip(field_names, row)) for row in DECISION_SPECS]


def validate_fixed_manifests(workbook_keys: set[str]) -> None:
    specs = _spec_rows()
    approved_keys = {row["source_event_key"] for row in specs}
    deferred_keys = set(DEFERRED_REASONS)
    if len(specs) != 55 or len(approved_keys) != 55:
        raise RuntimeError("fixed operator list must contain 55 unique events")
    if len(workbook_keys) != 63:
        raise RuntimeError("review workbook must contain 63 unique events")
    if _hash_lines(workbook_keys) != EXPECTED_ALL_EVENT_KEYS_HASH:
        raise RuntimeError("review workbook event-key manifest drifted")
    if approved_keys & deferred_keys:
        raise RuntimeError("approved and deferred event lists overlap")
    if approved_keys | deferred_keys != workbook_keys:
        raise RuntimeError("approved/deferred lists do not cover the workbook")
    if _hash_lines(approved_keys) != EXPECTED_APPROVED_EVENT_KEYS_HASH:
        raise RuntimeError("approved event-key manifest drifted")
    manifest = {
        f"{row['source_event_key']}|{row['expected_row_hash']}"
        for row in specs
    }
    if _hash_lines(manifest) != EXPECTED_OBSERVATION_MANIFEST_HASH:
        raise RuntimeError("CNInfo observation manifest drifted")


def _single_row(
    connection: sqlite3.Connection,
    query: str,
    params: tuple[Any, ...],
    message: str,
) -> sqlite3.Row:
    rows = connection.execute(query, params).fetchall()
    if len(rows) != 1:
        raise RuntimeError(message)
    return rows[0]


def _frozen_blocker_rows(
    connection: sqlite3.Connection,
    event_keys: Iterable[str],
) -> list[sqlite3.Row]:
    normalized_keys = sorted({
        str(event_key).strip()
        for event_key in event_keys
        if str(event_key).strip()
    })
    if not normalized_keys:
        return []
    placeholders = ",".join("?" for _ in normalized_keys)
    return connection.execute(
        f"""
        SELECT s.instrument_id, s.source_event_key, s.resolution_state
        FROM corporate_action_resolution_states AS s
        JOIN corporate_action_observations AS o
          ON o.source_event_key=s.source_event_key
         AND o.instrument_id=s.instrument_id
         AND o.source='cninfo'
         AND o.is_current=1
        WHERE s.factor_blocking=1
          AND s.source_event_key IN ({placeholders})
        ORDER BY s.instrument_id, s.source_event_key
        """,
        normalized_keys,
    ).fetchall()


def _notes(factor_effect: str, *, uses_tdx_date: bool) -> str:
    parts = [
        "用户确认该事项属于股改、重整、补偿、偿债或类似原因引发的非对称分派送转。",
        "保留CNInfo公司全体股东口径经济数字，不采用TDX经济数字或因子。",
    ]
    if factor_effect == "none":
        parts.append("该事项不对上市流通股东发生除权影响，事件保留但因子影响为0。")
    elif factor_effect == "official_reference_price":
        parts.append("因子仅采用CNInfo官方公告给出的调整前后参考价格之比。")
    if uses_tdx_date:
        parts.append("CNInfo日期与市场除权日冲突，仅采用指定TDX除权交易日。")
    return "".join(parts)


def build_decisions(
    connection: sqlite3.Connection,
    workbook_keys: set[str],
) -> list[dict[str, Any]]:
    validate_fixed_manifests(workbook_keys)
    decisions: list[dict[str, Any]] = []
    for spec in _spec_rows():
        event_key = spec["source_event_key"]
        instrument_id = spec["instrument_id"]
        observation = _single_row(
            connection,
            """
            SELECT *
            FROM corporate_action_observations
            WHERE source_event_key=? AND instrument_id=?
              AND source='cninfo' AND is_current=1
            """,
            (event_key, instrument_id),
            f"current CNInfo observation missing or ambiguous: {event_key}",
        )
        if observation["row_hash"] != spec["expected_row_hash"]:
            raise RuntimeError(f"CNInfo observation row hash drifted: {event_key}")
        _single_row(
            connection,
            """
            SELECT id
            FROM corporate_action_llm_analyses
            WHERE id=? AND source_event_key=? AND instrument_id=?
            """,
            (spec["analysis_id"], event_key, instrument_id),
            f"frozen analysis identity missing: {event_key}",
        )
        _single_row(
            connection,
            """
            SELECT id
            FROM corporate_action_effective_date_evidence
            WHERE source_event_key=? AND instrument_id=?
              AND announcement_id=?
              AND evidence_source='cninfo_announcement_metadata'
            """,
            (event_key, instrument_id, spec["announcement_id"]),
            f"frozen CNInfo announcement candidate missing: {event_key}",
        )
        terms = {
            field_name: observation[field_name]
            for field_name in ECONOMIC_FIELDS
            if observation[field_name] is not None
        }
        if not terms:
            raise RuntimeError(f"CNInfo economic terms are empty: {event_key}")
        tdx_record_id = TDX_DATE_ROWS.get(event_key)
        if tdx_record_id is not None:
            tdx_row = _single_row(
                connection,
                """
                SELECT id, instrument_id, ex_date
                FROM adjustment_factors_tdx
                WHERE id=? AND instrument_id=?
                """,
                (tdx_record_id, instrument_id),
                f"frozen TDX date row missing: {event_key}",
            )
            if str(tdx_row["ex_date"])[:10] != spec["effective_date"]:
                raise RuntimeError(f"frozen TDX ex-date drifted: {event_key}")
            exchange = "SSE" if instrument_id.endswith(".SH") else "SZSE"
            _single_row(
                connection,
                """
                SELECT id
                FROM trading_calendar
                WHERE exchange=? AND date(date)=? AND is_trading_day=1
                """,
                (exchange, spec["effective_date"]),
                f"TDX date is not an exchange trading session: {event_key}",
            )
        factor_reference = None
        if spec["factor_effect"] == "official_reference_price":
            before, after = OFFICIAL_FACTOR_REFERENCES[instrument_id]
            factor_reference = {
                "pre_adjustment_reference_price": before,
                "adjusted_reference_price": after,
            }
        payload = {
            **spec,
            "reviewer": REVIEWER,
            "approval_classification": "approved_asymmetric",
            "beneficiary_scope": "股改、重整、补偿、偿债等非对称事项的特定受益股东",
            "beneficiary_terms": {
                "economic_terms_source": "CNInfo公司全体股东口径",
                "asymmetric_beneficiary_terms_applied_to_factor": False,
            },
            "total_share_capital_terms": terms,
            "notes": _notes(
                spec["factor_effect"],
                uses_tdx_date=tdx_record_id is not None,
            ),
        }
        payload.pop("expected_row_hash")
        if factor_reference is not None:
            payload["factor_reference"] = factor_reference
        if tdx_record_id is not None:
            payload["tdx_record_id"] = tdx_record_id
            payload["expected_tdx_ex_date"] = spec["effective_date"]
        decisions.append(payload)
    decisions.sort(key=lambda row: (row["instrument_id"], row["source_event_key"]))
    return decisions


def validate_decision_payload(decisions: list[dict[str, Any]]) -> str:
    payload_hash = _canonical_hash(decisions)
    if (
        EXPECTED_DECISION_PAYLOAD_HASH
        and payload_hash != EXPECTED_DECISION_PAYLOAD_HASH
    ):
        raise RuntimeError("complete operator decision payload drifted")
    return payload_hash


def _hash_query(
    connection: sqlite3.Connection,
    query: str,
    params: tuple[Any, ...] = (),
) -> dict[str, Any]:
    rows = connection.execute(query, params).fetchall()
    normalized = [dict(row) for row in rows]
    return {"rows": len(normalized), "sha256": _canonical_hash(normalized)}


def immutable_snapshot(
    connection: sqlite3.Connection,
    decisions: list[dict[str, Any]],
) -> dict[str, dict[str, Any]]:
    instrument_ids = sorted({row["instrument_id"] for row in decisions})
    event_keys = sorted({row["source_event_key"] for row in decisions})
    instruments = ",".join("?" for _ in instrument_ids)
    events = ",".join("?" for _ in event_keys)
    snapshot = {
        "cninfo_observations": _hash_query(
            connection,
            f"""
            SELECT *
            FROM corporate_action_observations
            WHERE source='cninfo' AND source_event_key IN ({events})
            ORDER BY id
            """,
            tuple(event_keys),
        ),
        "tdx_rows": _hash_query(
            connection,
            f"""
            SELECT *
            FROM adjustment_factors_tdx
            WHERE instrument_id IN ({instruments})
            ORDER BY id
            """,
            tuple(instrument_ids),
        ),
    }
    for table in PRODUCTION_FACTOR_TABLES:
        snapshot[table] = _hash_query(
            connection,
            f"""
            SELECT *
            FROM {table}
            WHERE instrument_id IN ({instruments})
            ORDER BY id
            """,
            tuple(instrument_ids),
        )
    snapshot["adjustment_factor_series_status"] = _hash_query(
        connection,
        """
        SELECT *
        FROM adjustment_factor_series_status
        ORDER BY series_version
        """,
    )
    return snapshot


async def _apply_decisions(
    decisions: list[dict[str, Any]],
    *,
    database_path: Path,
) -> list[dict[str, Any]]:
    if database_path.resolve() != DEFAULT_DATABASE.resolve():
        raise ValueError(
            "--write is restricted to the project's configured quotes.db"
        )
    # The project configuration loader resolves its config directory from cwd.
    # Keep the CLI rooted for the lifetime of the shared database manager.
    os.chdir(ROOT_DIR)
    from data_manager import DataManager

    manager = DataManager()
    results = []
    for index, payload in enumerate(decisions, start=1):
        result = await manager.review_cninfo_asymmetric_manual_override(
            dict(payload)
        )
        review_id = result.get("review", {}).get("review_id")
        if not review_id:
            raise RuntimeError(
                f"review write did not return an identity: "
                f"{payload['source_event_key']}"
            )
        results.append({
            "sequence": index,
            "instrument_id": payload["instrument_id"],
            "source_event_key": payload["source_event_key"],
            "review_id": review_id,
            "factor_effect": result["factor_effect"],
            "factor_override": result["factor_override"],
            "tdx_date_used": result["tdx_date_used"],
            "tdx_economic_terms_used": result["tdx_economic_terms_used"],
            "tdx_factor_used": result["tdx_factor_used"],
        })
    return results


def audit_written_decisions(
    connection: sqlite3.Connection,
    decisions: list[dict[str, Any]],
) -> dict[str, Any]:
    event_keys = sorted({row["source_event_key"] for row in decisions})
    placeholders = ",".join("?" for _ in event_keys)
    latest_rows = connection.execute(
        f"""
        SELECT r.*
        FROM corporate_action_resolution_reviews AS r
        WHERE r.source_event_key IN ({placeholders})
          AND r.id = (
              SELECT MAX(latest.id)
              FROM corporate_action_resolution_reviews AS latest
              WHERE latest.source_event_key = r.source_event_key
          )
        ORDER BY r.source_event_key
        """,
        event_keys,
    ).fetchall()
    if len(latest_rows) != 55:
        raise RuntimeError("latest review audit did not find all 55 decisions")
    effect_counts = {
        "normal": 0,
        "none": 0,
        "official_reference_price": 0,
    }
    tdx_date_count = 0
    for row in latest_rows:
        if row["reviewer"] != REVIEWER or row["decision"] != "resolved":
            raise RuntimeError(
                f"latest review is not the fixed operator decision: "
                f"{row['source_event_key']}"
            )
        payload = json.loads(row["review_payload_json"])
        effect = str(payload.get("factor_effect") or "")
        if effect not in effect_counts:
            raise RuntimeError(f"unexpected factor effect: {effect}")
        effect_counts[effect] += 1
        if payload.get("tdx_date_used"):
            tdx_date_count += 1
            if (
                payload.get("tdx_economic_terms_used") is not False
                or payload.get("tdx_factor_used") is not False
            ):
                raise RuntimeError("TDX economics or factor leaked into CNInfo")
    if effect_counts != {
        "normal": 31,
        "none": 19,
        "official_reference_price": 5,
    }:
        raise RuntimeError(f"unexpected factor-effect counts: {effect_counts}")
    if tdx_date_count != 3:
        raise RuntimeError(f"unexpected TDX date-only count: {tdx_date_count}")
    blocker_rows = _frozen_blocker_rows(
        connection,
        set(event_keys) | set(DEFERRED_REASONS),
    )
    blocker_keys = {row["source_event_key"] for row in blocker_rows}
    if blocker_keys != set(DEFERRED_REASONS):
        raise RuntimeError(
            "post-write blockers do not equal the eight deferred events"
        )
    return {
        "review_count": len(latest_rows),
        "factor_effect_counts": effect_counts,
        "tdx_date_only_count": tdx_date_count,
        "remaining_blocker_count": len(blocker_rows),
        "remaining_blockers": [dict(row) for row in blocker_rows],
    }


def _summary(decisions: list[dict[str, Any]]) -> dict[str, Any]:
    counts = {
        "normal": 0,
        "none": 0,
        "official_reference_price": 0,
    }
    for row in decisions:
        counts[row["factor_effect"]] += 1
    return {
        "decision_count": len(decisions),
        "factor_effect_counts": counts,
        "tdx_date_only_count": sum(
            "tdx_record_id" in row for row in decisions
        ),
        "deferred_count": len(DEFERRED_REASONS),
        "deferred_events": DEFERRED_REASONS,
    }


def partial_apply_status(
    database_path: Path,
    decisions: list[dict[str, Any]],
    error: Exception,
) -> dict[str, Any]:
    connection = sqlite3.connect(database_path)
    try:
        rows = connection.execute(
            """
            SELECT DISTINCT source_event_key
            FROM corporate_action_resolution_reviews
            WHERE reviewer=?
            """,
            (REVIEWER,),
        ).fetchall()
    finally:
        connection.close()
    persisted_keys = {str(row[0]) for row in rows}
    expected_keys = {row["source_event_key"] for row in decisions}
    pending_keys = sorted(expected_keys - persisted_keys)
    return {
        "status": "write_or_audit_failed_rerun_required",
        "error_type": type(error).__name__,
        "error_message": str(error),
        "persisted_decision_count": len(expected_keys & persisted_keys),
        "pending_decision_count": len(pending_keys),
        "pending_event_keys": pending_keys,
        "resume": (
            "Rerun this fixed command with --write; review identities are "
            "idempotent and the complete post-write audit will run again."
        ),
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--write",
        action="store_true",
        help="Persist the exact 55 operator-approved review bundles.",
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=DEFAULT_DATABASE,
        help="SQLite database path.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="The frozen 63-event review workbook.",
    )
    args = parser.parse_args()

    workbook_keys = _workbook_event_keys(args.workbook)
    connection = sqlite3.connect(args.database)
    connection.row_factory = sqlite3.Row
    try:
        decisions = build_decisions(connection, workbook_keys)
        decision_hash = validate_decision_payload(decisions)
        before = immutable_snapshot(connection, decisions)
    finally:
        connection.close()

    result: dict[str, Any] = {
        "status": "validated_preview",
        "write_requested": bool(args.write),
        "decision_payload_hash": decision_hash,
        **_summary(decisions),
        "immutable_snapshot_before": before,
    }
    if args.write:
        try:
            result["writes"] = asyncio.run(_apply_decisions(
                decisions,
                database_path=args.database,
            ))
            connection = sqlite3.connect(args.database)
            connection.row_factory = sqlite3.Row
            try:
                after = immutable_snapshot(connection, decisions)
                if after != before:
                    raise RuntimeError(
                        "raw CNInfo, TDX, or production-factor rows changed"
                    )
                result["audit"] = audit_written_decisions(
                    connection,
                    decisions,
                )
            finally:
                connection.close()
            result["immutable_snapshot_after"] = after
            result["status"] = "applied_and_audited"
        except Exception as error:
            result.update(partial_apply_status(
                args.database,
                decisions,
                error,
            ))
            print(json.dumps(
                result,
                ensure_ascii=False,
                indent=2,
                default=str,
            ))
            return 2
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
